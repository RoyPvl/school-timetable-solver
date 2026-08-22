from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from typing import Any

from openpyxl.utils.cell import range_boundaries

from school_timetable_solver.adapter.excel_input_v2_adapter import BLANK_TOKEN, UNSPECIFIED
from school_timetable_solver.adapter.excel_input_v2_reference_adapter import (
    ReferenceLabelExcelInputV2ReaderAdapter,
)
from school_timetable_solver.model.master_models import ClassModel
from school_timetable_solver.model.result_models import InputReadResultModel, ValidationIssueModel


class EntityReferenceExcelInputV2ReaderAdapter(ReferenceLabelExcelInputV2ReaderAdapter):
    """Read v2 references for teachers, subjects, and classes.

    Class labels are unique within a campus, not globally. During parsing the
    temporary ``ClassModel.class_name`` is the input label so the existing
    campus+class resolver remains unambiguous. Before returning ``InputDataModel``
    the original output display name is restored from ``T_REFERENCE_MAP``.
    """

    def __init__(self) -> None:
        super().__init__()
        self._class_label_by_id: dict[str, str] = {}
        self._class_output_name_by_id: dict[str, str] = {}

    def read(self, path: Path) -> InputReadResultModel:
        result = super().read(path)
        if result.input_data is None or not self._class_output_name_by_id:
            return result
        restored_classes = tuple(
            replace(
                item,
                class_name=self._class_output_name_by_id.get(item.class_id, item.class_name),
            )
            for item in result.input_data.classes
        )
        return InputReadResultModel(
            replace(result.input_data, classes=restored_classes),
            result.issues,
        )

    def _reset_reference_maps(self) -> None:
        super()._reset_reference_maps()
        self._class_label_by_id = {}
        self._class_output_name_by_id = {}

    def _load_reference_map(self, workbook: Any, issues: list[ValidationIssueModel]) -> None:
        table = None
        table_sheet = None
        for worksheet in workbook.worksheets:
            if "T_REFERENCE_MAP" in worksheet.tables:
                table = worksheet.tables["T_REFERENCE_MAP"]
                table_sheet = worksheet
                break
        if table is None or table_sheet is None:
            return

        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = {
            str(table_sheet.cell(min_row, column).value).strip(): column
            for column in range(min_col, max_col + 1)
        }
        required = {"参照種別", "内部ID", "入力用ラベル", "出力表示名"}
        missing = required.difference(headers)
        if missing:
            issues.append(
                self._issue(
                    "V2_REFERENCE_MAP_HEADER",
                    "T_REFERENCE_MAP",
                    f"参照マップの必須列がありません: {', '.join(sorted(missing))}",
                )
            )
            return

        teacher_labels: dict[str, str] = {}
        subject_labels: dict[str, str] = {}
        teacher_names: dict[str, str] = {}
        subject_names: dict[str, str] = {}
        class_labels: dict[str, str] = {}
        class_names: dict[str, str] = {}

        for row_index in range(min_row + 1, max_row + 1):
            entity = self._optional_text(table_sheet.cell(row_index, headers["参照種別"]).value)
            internal_id = self._optional_text(table_sheet.cell(row_index, headers["内部ID"]).value)
            label = self._optional_text(table_sheet.cell(row_index, headers["入力用ラベル"]).value)
            raw_output = table_sheet.cell(row_index, headers["出力表示名"]).value
            output_name = "" if raw_output is None else str(raw_output)
            if entity is None and internal_id is None and label is None:
                continue
            if entity not in {"teacher", "subject", "class"} or internal_id is None or label is None:
                issues.append(
                    self._issue(
                        "V2_REFERENCE_MAP_ROW",
                        f"T_REFERENCE_MAP!{row_index}",
                        "参照種別・内部ID・入力用ラベルを正しく入力してください",
                    )
                )
                continue

            if entity == "teacher":
                if label in teacher_labels:
                    issues.append(self._issue("V2_DUPLICATE_INPUT_LABEL", label, f"教師の入力用ラベルが重複しています: {label}"))
                    continue
                if internal_id in teacher_names:
                    issues.append(self._issue("V2_DUPLICATE_REFERENCE_ID", internal_id, f"教師の内部IDが重複しています: {internal_id}"))
                    continue
                teacher_labels[label] = internal_id
                teacher_names[internal_id] = output_name
            elif entity == "subject":
                if label in subject_labels:
                    issues.append(self._issue("V2_DUPLICATE_INPUT_LABEL", label, f"教科の入力用ラベルが重複しています: {label}"))
                    continue
                if internal_id in subject_names:
                    issues.append(self._issue("V2_DUPLICATE_REFERENCE_ID", internal_id, f"教科の内部IDが重複しています: {internal_id}"))
                    continue
                subject_labels[label] = internal_id
                subject_names[internal_id] = output_name
            else:
                if internal_id in class_labels:
                    issues.append(self._issue("V2_DUPLICATE_REFERENCE_ID", internal_id, f"クラスの内部IDが重複しています: {internal_id}"))
                    continue
                class_labels[internal_id] = label
                class_names[internal_id] = output_name

        self._teacher_label_to_id = teacher_labels
        self._subject_label_to_id = subject_labels
        self._teacher_output_name_by_id = teacher_names
        self._subject_output_name_by_id = subject_names
        self._class_label_by_id = class_labels
        self._class_output_name_by_id = class_names

    def _read_classes(
        self,
        rows: list[dict[str, Any]],
        campus_by_name: dict[str, str],
        teacher_by_name: dict[str, str],
        issues: list[ValidationIssueModel],
    ) -> list[ClassModel]:
        if not self._class_label_by_id:
            return super()._read_classes(rows, campus_by_name, teacher_by_name, issues)

        result: list[ClassModel] = []
        seen_keys: set[tuple[str, str]] = set()
        for index, row in enumerate(rows, start=1):
            internal_id = self._required_text(row.get("内部ID"), "T_CLASSES", f"行{index}/内部ID", issues)
            campus_id = self._resolve_name(row.get("校舎"), campus_by_name, "校舎", "T_CLASSES", index, issues)
            input_label = self._required_text(row.get("クラス"), "T_CLASSES", f"行{index}/クラス", issues)
            division_display = self._required_text(row.get("学部"), "T_CLASSES", f"行{index}", issues)
            division = self._division_to_internal.get(
                division_display or "",
                self._optional_text(row.get("内部学部")) or division_display,
            )
            grade = self._required_int(row.get("学年"), "T_CLASSES", f"行{index}", issues)
            exam_category = self._required_text(row.get("区分"), "T_CLASSES", f"行{index}", issues)
            homeroom_display = self._optional_text(row.get("担任"))
            homeroom_id = (
                None
                if homeroom_display in (None, UNSPECIFIED, BLANK_TOKEN)
                else self._resolve_name(
                    homeroom_display,
                    teacher_by_name,
                    "教師",
                    "T_CLASSES",
                    index,
                    issues,
                )
            )
            enabled = self._required_yes_no(row.get("使用"), "T_CLASSES", f"行{index}", issues)
            if None in (internal_id, campus_id, input_label, division, grade, exam_category, enabled):
                continue
            expected_label = self._class_label_by_id.get(str(internal_id))
            if expected_label != input_label:
                issues.append(
                    self._issue(
                        "V2_REFERENCE_LABEL_MISMATCH",
                        input_label,
                        f"クラスの入力用ラベルと内部IDの対応が参照マップと一致しません: {input_label}",
                    )
                )
                continue
            key = (str(campus_id), input_label)
            if key in seen_keys:
                issues.append(
                    self._issue(
                        "V2_DUPLICATE_CLASS_INPUT_LABEL",
                        f"{campus_id}/{input_label}",
                        f"同一校舎でクラス入力用ラベルが重複しています: {input_label}",
                    )
                )
                continue
            seen_keys.add(key)
            result.append(
                ClassModel(
                    str(internal_id),
                    input_label,
                    str(campus_id),
                    str(division),
                    int(grade),
                    str(exam_category),
                    homeroom_id,
                    bool(enabled),
                )
            )
        return result
