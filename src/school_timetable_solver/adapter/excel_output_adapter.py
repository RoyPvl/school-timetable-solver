from __future__ import annotations

import os
import tempfile
from datetime import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from school_timetable_solver.model.result_models import TimetableDocumentModel

_THIN = Side(style="thin")
_HAIR = Side(style="hair")
_WEEKDAYS = ("月曜日", "火曜日", "水曜日", "木曜日", "金曜日", "土曜日", "日曜日")


class ExcelTimetableWriterAdapter:
    """Write one-sheet date matrices and atomically replace the destination."""

    def write(self, document: TimetableDocumentModel, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = "全体"
        room_count = sum(len(campus.rooms) for campus in document.campuses)
        matrix_width = 2 + room_count
        for date_index, _daily in enumerate(document.dates):
            start_row = 2 + (date_index // 2) * 22
            start_column = 1 if date_index % 2 == 0 else 1 + matrix_width + 1
            self._write_daily_matrix(
                worksheet,
                document,
                date_index,
                start_row,
                start_column,
            )
        self._atomic_save(workbook, path)

    def _write_daily_matrix(
        self,
        worksheet: Worksheet,
        document: TimetableDocumentModel,
        date_index: int,
        start_row: int,
        start_column: int,
    ) -> None:
        daily = document.dates[date_index]
        flat_rooms = [
            (campus.campus_id, room) for campus in document.campuses for room in campus.rooms
        ]
        worksheet.cell(start_row, start_column, daily.target_date)
        worksheet.cell(start_row, start_column).number_format = 'm"月"d"日"'
        worksheet.cell(start_row + 1, start_column, _WEEKDAYS[daily.target_date.weekday()])

        room_offset = 0
        for campus in document.campuses:
            campus_column = start_column + 2 + room_offset
            if campus.rooms:
                worksheet.cell(start_row, campus_column, campus.campus_display_name)
                if len(campus.rooms) > 1:
                    worksheet.merge_cells(
                        start_row=start_row,
                        start_column=campus_column,
                        end_row=start_row,
                        end_column=campus_column + len(campus.rooms) - 1,
                    )
                for index, room in enumerate(campus.rooms):
                    worksheet.cell(
                        start_row + 1,
                        campus_column + index,
                        room.room_display_name,
                    )
            room_offset += len(campus.rooms)

        for period_index, period in enumerate(document.periods):
            class_row = start_row + 2 + period_index * 3
            worksheet.cell(
                class_row,
                start_column,
                (
                    f"{period.period_name}\n"
                    f"{self._format_time(period.start_time)}～"  # noqa: RUF001
                    f"{self._format_time(period.end_time)}"
                ),
            )
            worksheet.cell(class_row, start_column).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            worksheet.cell(class_row, start_column + 1, "クラス")
            worksheet.cell(class_row + 1, start_column + 1, "教科")
            worksheet.cell(class_row + 2, start_column + 1, "担当")
            for room_index, (_, room) in enumerate(flat_rooms):
                lesson = daily.lessons_by_period_and_room.get((period.period_id, room.room_id))
                if lesson is None:
                    continue
                column = start_column + 2 + room_index
                worksheet.cell(class_row, column, lesson.class_display_name)
                worksheet.cell(class_row + 1, column, lesson.subject_display_name)
                worksheet.cell(class_row + 2, column, lesson.teacher_display_name)

        matrix_width = 2 + len(flat_rooms)
        self._apply_matrix_borders(
            worksheet,
            start_row,
            start_column,
            matrix_width,
            tuple(campus_id for campus_id, _ in flat_rooms),
        )
        worksheet.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=start_row,
            end_column=start_column + 1,
        )
        worksheet.merge_cells(
            start_row=start_row + 1,
            start_column=start_column,
            end_row=start_row + 1,
            end_column=start_column + 1,
        )
        for period_index in range(6):
            class_row = start_row + 2 + period_index * 3
            worksheet.merge_cells(
                start_row=class_row,
                start_column=start_column,
                end_row=class_row + 2,
                end_column=start_column,
            )
        self._apply_merged_cell_borders(
            worksheet,
            document,
            start_row,
            start_column,
        )
        self._write_teacher_leaves(
            worksheet,
            document,
            date_index,
            start_row,
            start_column,
        )

    def _write_teacher_leaves(
        self,
        worksheet: Worksheet,
        document: TimetableDocumentModel,
        date_index: int,
        start_row: int,
        start_column: int,
    ) -> None:
        leave_row = start_row + 20
        daily = document.dates[date_index]
        all_period_ids = {period.period_id for period in document.periods}
        room_offset = 0
        for campus in document.campuses:
            cursor = start_column + 2 + room_offset
            campus_end = cursor + len(campus.rooms) - 1
            for teacher_leave in (
                item for item in daily.teacher_leaves if item.campus_id == campus.campus_id
            ):
                is_full_day = set(teacher_leave.unavailable_period_ids) == all_period_ids
                required_cells = 1 if is_full_day else 2
                if cursor + required_cells - 1 > campus_end:
                    raise ValueError(
                        f"OUTPUT_TEACHER_LEAVE_OVERFLOW: {daily.target_date}/{campus.campus_id}"
                    )
                name_cell = worksheet.cell(
                    leave_row,
                    cursor,
                    teacher_leave.teacher_display_name,
                )
                self._style_teacher_leave_cell(name_cell)
                if not is_full_day:
                    period_cell = worksheet.cell(
                        leave_row,
                        cursor + 1,
                        self._format_teacher_leave_periods(
                            teacher_leave.unavailable_period_ids,
                            document,
                        ),
                    )
                    self._style_teacher_leave_cell(period_cell)
                cursor += required_cells
            room_offset += len(campus.rooms)

    def _style_teacher_leave_cell(self, cell: Cell) -> None:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _format_teacher_leave_periods(
        self,
        unavailable_period_ids: tuple[str, ...],
        document: TimetableDocumentModel,
    ) -> str:
        period_positions = {
            period.period_id: index for index, period in enumerate(document.periods)
        }
        period_names = {period.period_id: period.period_name for period in document.periods}
        sorted_ids = sorted(
            unavailable_period_ids,
            key=period_positions.__getitem__,
        )
        runs: list[list[str]] = []
        for period_id in sorted_ids:
            if not runs or period_positions[period_id] != period_positions[runs[-1][-1]] + 1:
                runs.append([period_id])
            else:
                runs[-1].append(period_id)
        parts: list[str] = []
        for run in runs:
            if len(run) >= 3:
                parts.append(f"{period_names[run[0]]}～{period_names[run[-1]]}")  # noqa: RUF001
            else:
                parts.append("".join(period_names[period_id] for period_id in run))
        return "".join(parts)

    def _apply_merged_cell_borders(
        self,
        worksheet: Worksheet,
        document: TimetableDocumentModel,
        start_row: int,
        start_column: int,
    ) -> None:
        worksheet.cell(start_row, start_column).border = Border(
            left=_THIN,
            right=_THIN,
            top=_THIN,
            bottom=_HAIR,
        )
        worksheet.cell(start_row + 1, start_column).border = Border(
            left=_THIN,
            right=_THIN,
            top=_HAIR,
            bottom=_THIN,
        )
        room_offset = 0
        for campus in document.campuses:
            if len(campus.rooms) > 1:
                worksheet.cell(
                    start_row,
                    start_column + 2 + room_offset,
                ).border = Border(
                    left=_THIN,
                    right=_THIN,
                    top=_THIN,
                    bottom=_HAIR,
                )
            room_offset += len(campus.rooms)
        for period_index in range(6):
            worksheet.cell(
                start_row + 2 + period_index * 3,
                start_column,
            ).border = Border(
                left=_THIN,
                right=_HAIR,
                top=_THIN,
                bottom=_THIN,
            )

    def _apply_matrix_borders(
        self,
        worksheet: Worksheet,
        start_row: int,
        start_column: int,
        matrix_width: int,
        room_campus_ids: tuple[str, ...],
    ) -> None:
        vertical_boundaries = [_THIN, _HAIR, _THIN]
        for index in range(1, len(room_campus_ids)):
            vertical_boundaries.append(
                _THIN if room_campus_ids[index] != room_campus_ids[index - 1] else _HAIR
            )
        vertical_boundaries.append(_THIN)
        horizontal_boundaries = [_THIN, _HAIR, _THIN]
        for _ in range(6):
            horizontal_boundaries.extend((_HAIR, _HAIR, _THIN))
        for row_offset in range(20):
            for column_offset in range(matrix_width):
                cell = worksheet.cell(
                    start_row + row_offset,
                    start_column + column_offset,
                )
                cell.border = Border(
                    left=vertical_boundaries[column_offset],
                    right=vertical_boundaries[column_offset + 1],
                    top=horizontal_boundaries[row_offset],
                    bottom=horizontal_boundaries[row_offset + 1],
                )

    def _format_time(self, value: time) -> str:
        return f"{value.hour}:{value.minute:02d}"

    def _atomic_save(self, workbook: Workbook, path: Path) -> None:
        descriptor, temp_name = tempfile.mkstemp(
            prefix=f".{path.stem}-",
            suffix=".xlsx",
            dir=path.parent,
        )
        os.close(descriptor)
        temp_path = Path(temp_name)
        completed = False
        try:
            workbook.save(temp_path)
            temp_path.replace(path)
            completed = True
        finally:
            if not completed and temp_path.exists():
                temp_path.unlink()
