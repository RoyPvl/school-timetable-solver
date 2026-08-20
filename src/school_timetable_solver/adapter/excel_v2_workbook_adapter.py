from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from school_timetable_solver.adapter.excel_input_v2_adapter import BLANK_TOKEN, NO, UNSPECIFIED, YES
from school_timetable_solver.model.input_models import InputDataModel


class ExcelV2WorkbookWriterAdapter:
    """Write the v2.0 human/AI self-describing workbook from existing input models."""

    INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
    AUTO_FILL = PatternFill("solid", fgColor="E7E6E6")
    TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
    HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
    ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")

    def write(self, path: Path, data: InputDataModel) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_intro(workbook)
        self._write_schedule(workbook, data)
        self._write_campuses_rooms(workbook, data)
        self._write_teachers(workbook, data)
        self._write_classes_lessons(workbook, data)
        self._write_teacher_conditions(workbook, data)
        self._write_placement_rules(workbook, data)
        self._write_individual_rules(workbook, data)
        self._write_check_sheet(workbook, data)
        self._write_ai_guide(workbook)
        self._write_system(workbook, data)
        workbook["_system"].sheet_state = "hidden"
        workbook.active = workbook.sheetnames.index("00_最初に読む")
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    def _write_intro(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("00_最初に読む")
        sheet["A1"] = "時間割入力Excel v2.0"
        sheet["A1"].font = Font(size=18, bold=True)
        sheet["A3"] = "このExcelは、時間割生成システムへ渡す入力ファイルです。Excelに詳しくない方とAIの双方が安全に編集できるように作られています。"
        sheet["A5"] = "入力の基本ルール"
        sheet["A5"].font = Font(size=14, bold=True)
        rules = (
            "○は「はい・有効・対象」、×は「いいえ・無効・対象外」です。必ずプルダウンから選択してください。",
            f"{BLANK_TOKEN} は意図的な空欄値です。実際にセルを空にすることとは意味が違います。",
            f"{UNSPECIFIED} は「この項目では条件を設定しない」という意味です。×とは異なります。",
            "日付の集合は、開始日～終了日の範囲へ勝手に置き換えないでください。除外されている日には意味があります。",
            "クラスや教室は同名が存在し得るため、必ず校舎と組み合わせて識別してください。",
            "黄色のセルは入力欄、灰色のセルは自動・内部情報です。色だけでなく列名と説明も確認してください。",
            "編集後は 08_入力確認 を確認してください。",
        )
        for row, text in enumerate(rules, start=6):
            sheet.cell(row, 1, f"・{text}")
        sheet["A15"] = "AIが編集する場合"
        sheet["A15"].font = Font(size=14, bold=True)
        ai_rules = (
            "最初にこのシートと 90_AI編集ガイド を読むこと。",
            "利用者の指示にない条件を推測して追加しないこと。",
            "固定された対象集合を、学年・学部等の動的な条件へ一般化しないこと。",
            f"実セルの空欄と {BLANK_TOKEN} を置換しないこと。",
            "_system シートや『内部』で始まる列は原則変更しないこと。",
            "判断できない場合は推測せず利用者へ確認すること。",
        )
        for row, text in enumerate(ai_rules, start=16):
            sheet.cell(row, 1, f"・{text}")
        sheet.column_dimensions["A"].width = 120
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "A5"

    def _write_schedule(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("01_日程")
        row = self._section(sheet, 1, "基本情報", "時間割名と説明です。schema_versionはシステム側で管理します。")
        row = self._add_table(sheet, row, "T_SETTINGS", ["項目", "値"], [["時間割名", data.settings.timetable_name], ["説明", data.settings.description or ""]]) + 2
        row = self._section(sheet, row, "時限", "時限IDは内部情報です。通常は『時限』『開始』『終了』だけ確認してください。")
        period_rows = [[p.period_id, p.period_name, p.output_order, p.start_time, p.end_time] for p in sorted(data.periods, key=lambda item: item.output_order)]
        row = self._add_table(sheet, row, "T_PERIODS", ["内部ID", "時限", "表示順", "開始", "終了"], period_rows, hidden_columns={"内部ID"}) + 2
        row = self._section(sheet, row, "開講カレンダー", "1日1行です。日付範囲へまとめないでください。○の時限だけ授業配置可能です。")
        ordered_periods = sorted(data.periods, key=lambda item: item.output_order)
        headers = ["日付", "出力", *[f"{i}限" for i in range(1, len(ordered_periods) + 1)], "備考"]
        calendar_rows = []
        for day in sorted(data.calendar_days, key=lambda item: item.target_date):
            enabled = set(day.enabled_period_ids)
            calendar_rows.append([day.target_date, self._yn(day.output_enabled), *[self._yn(p.period_id in enabled) for p in ordered_periods], day.note or ""])
        row = self._add_table(sheet, row, "T_CALENDAR", headers, calendar_rows, yes_no_columns=set(headers[1:-1])) + 2
        row = self._section(sheet, row, "名前付き日付集合", "教師の休み候補日など、明示的な日付集合です。期間として解釈しません。")
        date_sets = self._date_sets(data)
        date_rows = [[name, target_date, ""] for name, dates in date_sets.items() for target_date in dates]
        self._add_table(sheet, row, "T_DATE_SETS", ["日付集合", "日付", "備考"], date_rows)
        self._finalize_sheet(sheet)

    def _write_campuses_rooms(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("02_校舎・教室")
        row = self._section(sheet, 1, "校舎", "校舎の一覧です。内部IDは編集しないでください。")
        campus_rows = [[c.campus_id, c.campus_name, c.output_order, self._yn(c.enabled), ""] for c in data.campuses]
        row = self._add_table(sheet, row, "T_CAMPUSES", ["内部ID", "校舎", "表示順", "使用", "備考"], campus_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"}) + 2
        row = self._section(sheet, row, "教室", "同名教室が別校舎に存在し得ます。必ず校舎と教室の組で扱います。")
        campus_names = {c.campus_id: c.campus_name for c in data.campuses}
        room_rows = [[r.room_id, campus_names[r.campus_id], r.room_name, r.output_order, r.priority, self._yn(r.enabled), ""] for r in data.rooms]
        self._add_table(sheet, row, "T_ROOMS", ["内部ID", "校舎", "教室", "表示順", "優先度", "使用", "備考"], room_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"})
        self._finalize_sheet(sheet)

    def _write_teachers(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("03_教師")
        row = self._section(sheet, 1, "教師", f"教師名が意図的に空欄の場合は {BLANK_TOKEN} と表示します。空セルとは別物です。")
        campus_names = {c.campus_id: c.campus_name for c in data.campuses}
        rows = []
        for teacher in data.teachers:
            display = teacher.teacher_name if teacher.teacher_name else BLANK_TOKEN
            rows.append([teacher.teacher_id, display, teacher.teacher_name, campus_names[teacher.home_campus_id], self._yn(teacher.enabled), ""])
        self._add_table(sheet, row, "T_TEACHERS", ["内部ID", "教師", "内部表示名", "主校舎", "使用", "備考"], rows, yes_no_columns={"使用"}, hidden_columns={"内部ID", "内部表示名"})
        self._finalize_sheet(sheet)

    def _write_classes_lessons(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("04_クラス・授業")
        campus_names = {c.campus_id: c.campus_name for c in data.campuses}
        teacher_names = {t.teacher_id: (t.teacher_name or BLANK_TOKEN) for t in data.teachers}
        class_names = {c.class_id: (campus_names[c.campus_id], c.class_name) for c in data.classes}
        subject_names = {s.subject_id: (s.subject_name or BLANK_TOKEN) for s in data.subjects}
        division_display = {"elementary": "小学", "junior_high": "中学", "high_school": "高校", "other": "その他"}
        lesson_type_display = {"regular": "通常", "special": "特別", "other": "その他"}

        row = self._section(sheet, 1, "クラス", "学部・学年・区分はルール判定に使います。クラス名から自動推定しません。")
        class_rows = []
        for item in data.classes:
            class_rows.append([item.class_id, campus_names[item.campus_id], item.class_name, division_display.get(item.division, item.division), item.division, item.grade, item.exam_category, teacher_names.get(item.homeroom_teacher_id, BLANK_TOKEN), self._yn(item.enabled), ""])
        row = self._add_table(sheet, row, "T_CLASSES", ["内部ID", "校舎", "クラス", "学部", "内部学部", "学年", "区分", "担任", "使用", "備考"], class_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID", "内部学部"}) + 2

        row = self._section(sheet, row, "教科", "種別『通常』は担任授業判定にも使われます。")
        subject_rows = []
        for item in data.subjects:
            display = item.subject_name or BLANK_TOKEN
            subject_rows.append([item.subject_id, display, item.subject_name, lesson_type_display.get(item.lesson_type, item.lesson_type), item.lesson_type, self._yn(item.enabled), ""])
        row = self._add_table(sheet, row, "T_SUBJECTS", ["内部ID", "教科", "内部表示名", "種別", "内部種別", "使用", "備考"], subject_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID", "内部表示名", "内部種別"}) + 2

        row = self._section(sheet, row, "授業要求", "1行が1つの『クラス×教科×担当教師』です。1クラス1行にまとめません。")
        req_rows = []
        for req in data.lesson_requirements:
            campus, class_name = class_names[req.class_id]
            req_rows.append([req.requirement_id, campus, class_name, subject_names[req.subject_id], teacher_names[req.teacher_id], req.required_periods, req.max_periods_per_day if req.max_periods_per_day is not None else UNSPECIFIED, self._yn(req.enabled), ""])
        self._add_table(sheet, row, "T_LESSON_REQUIREMENTS", ["内部ID", "校舎", "クラス", "教科", "教師", "必要コマ", "1日最大", "使用", "備考"], req_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"})
        self._finalize_sheet(sheet)

    def _write_teacher_conditions(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("05_教師条件")
        teacher_names = {t.teacher_id: (t.teacher_name or BLANK_TOKEN) for t in data.teachers}
        ordered_periods = sorted(data.periods, key=lambda item: item.output_order)
        row = self._section(sheet, 1, "教師の個別休み", "○を付けた時限だけ担当不可です。『1日休み』等へ勝手にまとめないでください。")
        headers = ["教師", "日付", *[f"{i}限休み" for i in range(1, len(ordered_periods) + 1)], "備考"]
        leave_rows = []
        for leave in data.teacher_leaves:
            unavailable = set(leave.unavailable_period_ids)
            leave_rows.append([teacher_names[leave.teacher_id], leave.target_date, *[self._yn(p.period_id in unavailable) for p in ordered_periods], ""])
        row = self._add_table(sheet, row, "T_TEACHER_LEAVES", headers, leave_rows, yes_no_columns=set(headers[2:-1])) + 2

        row = self._section(sheet, row, "教師の休日日数", "候補日は 01_日程 の名前付き日付集合を参照します。候補日集合を日付範囲に置き換えないでください。")
        date_sets = self._date_sets(data)
        set_name_by_dates = {dates: name for name, dates in date_sets.items()}
        day_rows = []
        for rule in data.teacher_day_off_rules:
            day_rows.append([rule.rule_id, teacher_names[rule.teacher_id], set_name_by_dates[rule.eligible_dates], self._display_optional(rule.required_days_off), self._display_optional(rule.minimum_days_off), self._display_optional(rule.maximum_days_off), self._display_optional(rule.preferred_days_off), rule.quota_group_id or UNSPECIFIED, self._display_optional(rule.group_required_days_off), self._yn(rule.enabled), ""])
        self._add_table(sheet, row, "T_TEACHER_DAY_OFF_RULES", ["内部ID", "教師", "候補日集合", "必須休み日数", "最低休み日数", "最大休み日数", "希望休み日数", "グループ", "グループ必須休み日数", "使用", "備考"], day_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"})
        self._finalize_sheet(sheet)

    def _write_placement_rules(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("06_基本配置ルール")
        campus_names = {c.campus_id: c.campus_name for c in data.campuses}
        class_lookup = {c.class_id: (campus_names[c.campus_id], c.class_name) for c in data.classes}
        teacher_names = {t.teacher_id: (t.teacher_name or BLANK_TOKEN) for t in data.teachers}
        ordered_periods = sorted(data.periods, key=lambda item: item.output_order)
        row = self._section(sheet, 1, "基本配置ルール", "追加制約は既存条件との積み重ね、上書きはその項目を置き換えます。意味が異なるため変更時は注意してください。")
        headers = ["内部ID", "ルール", "使用", "対象", "種別", "校舎", "クラス", "教師", "開始日", "終了日", "曜日", "時限を指定", *[f"{i}限許可" for i in range(1, len(ordered_periods) + 1)], "1日上限", "1限6限同日禁止", "連続登校上限", "希望連続登校上限", *[f"{i}限必須" for i in range(1, len(ordered_periods) + 1)], "優先度", "備考"]
        rows = []
        condition_rows: list[list[Any]] = []
        field_display = {"class_id": "クラス", "division": "学部", "grade": "学年", "exam_category": "受験区分", "campus_id": "校舎", "teacher_id": "教師"}
        operator_display = {"eq": "=", "ne": "≠", "in": "いずれか", "ge": "以上", "le": "以下", "between": "範囲"}
        for rule in data.placement_rules:
            direct_class: str | None = None
            direct_teacher: str | None = None
            external_conditions: list[tuple[str, str, str]] = []
            for field, operator, value in zip(rule.condition_fields, rule.condition_operators, rule.condition_values, strict=True):
                if field == "class_id" and operator == "eq" and value in class_lookup:
                    direct_class = class_lookup[value][1]
                elif field == "teacher_id" and operator == "eq" and value in teacher_names:
                    direct_teacher = teacher_names[value]
                else:
                    external_conditions.append((field, operator, value))
            allowed = set(rule.allowed_period_ids)
            required = set(rule.required_lesson_period_ids)
            rows.append([rule.rule_id, rule.rule_name, self._yn(rule.enabled), "クラス" if rule.target_entity == "class" else "教師", "上書き" if rule.constraint_type == "override" else "追加制約", campus_names.get(rule.campus_id, UNSPECIFIED), direct_class or UNSPECIFIED, direct_teacher or UNSPECIFIED, rule.start_date or UNSPECIFIED, rule.end_date or UNSPECIFIED, "/".join(rule.weekdays) if rule.weekdays else UNSPECIFIED, self._yn(bool(rule.allowed_period_ids)), *[self._yn(p.period_id in allowed) for p in ordered_periods], self._display_optional(rule.daily_hard_limit), self._display_optional_bool(rule.forbid_first_last_same_day), self._display_optional(rule.attendance_streak_limit), self._display_optional(rule.preferred_attendance_streak_limit), *[self._display_optional_bool(True if p.period_id in required else None) for p in ordered_periods], rule.priority, ""])
            for field, operator, value in external_conditions:
                condition_rows.append([rule.rule_name, field_display.get(field, field), field, operator_display.get(operator, operator), operator, value])
        yes_no_columns = {"使用", "時限を指定", *[f"{i}限許可" for i in range(1, len(ordered_periods) + 1)]}
        optional_yes_no_columns = {"1限6限同日禁止", *[f"{i}限必須" for i in range(1, len(ordered_periods) + 1)]}
        row = self._add_table(sheet, row, "T_PLACEMENT_RULES", headers, rows, yes_no_columns=yes_no_columns, optional_yes_no_columns=optional_yes_no_columns, hidden_columns={"内部ID"}) + 2
        row = self._section(sheet, row, "対象条件", "複数行ある場合はAND条件です。内部条件項目・内部比較は自動列です。")
        self._add_table(sheet, row, "T_PLACEMENT_RULE_CONDITIONS", ["ルール", "条件項目", "内部条件項目", "比較", "内部比較", "値"], condition_rows, hidden_columns={"内部条件項目", "内部比較"})
        self._finalize_sheet(sheet)

    def _write_individual_rules(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("07_個別ルール")
        campus_names = {c.campus_id: c.campus_name for c in data.campuses}
        class_lookup = {c.class_id: (campus_names[c.campus_id], c.class_name) for c in data.classes}
        subject_names = {s.subject_id: (s.subject_name or BLANK_TOKEN) for s in data.subjects}
        ordered_periods = sorted(data.periods, key=lambda item: item.output_order)

        row = self._section(sheet, 1, "授業配置数ルール（Hard）", "同じ内容を持つルール本体をまとめ、対象クラス・教科は別表で明示します。対象を学年条件等へ一般化しません。")
        hard_rows, hard_targets = self._group_lesson_count_rules(data.lesson_count_rule_segments, class_lookup, subject_names, ordered_periods, soft=False)
        hard_headers = ["ルール", "Segment", "使用", "配置数", "開始日", "終了日", *[f"{i}限対象" for i in range(1, len(ordered_periods) + 1)], "備考"]
        row = self._add_table(sheet, row, "T_LESSON_COUNT_HARD", hard_headers, hard_rows, yes_no_columns={"使用", *[f"{i}限対象" for i in range(1, len(ordered_periods) + 1)]}) + 2
        row = self._section(sheet, row, "Hard対象", "ここに列挙されたクラス×教科だけが対象です。")
        row = self._add_table(sheet, row, "T_LESSON_COUNT_HARD_TARGETS", ["ルール", "校舎", "クラス", "教科"], hard_targets) + 2

        row = self._section(sheet, row, "授業配置数ルール（Soft）", "希望配置数です。対象は別表で明示します。")
        soft_rows, soft_targets = self._group_lesson_count_rules(data.lesson_count_preference_rule_segments, class_lookup, subject_names, ordered_periods, soft=True)
        soft_headers = ["ルール", "Segment", "使用", "希望配置数", "開始日", "終了日", *[f"{i}限対象" for i in range(1, len(ordered_periods) + 1)], "備考"]
        row = self._add_table(sheet, row, "T_LESSON_COUNT_SOFT", soft_headers, soft_rows, yes_no_columns={"使用", *[f"{i}限対象" for i in range(1, len(ordered_periods) + 1)]}) + 2
        row = self._section(sheet, row, "Soft対象", "ここに列挙されたクラス×教科だけが対象です。")
        row = self._add_table(sheet, row, "T_LESSON_COUNT_SOFT_TARGETS", ["ルール", "校舎", "クラス", "教科"], soft_targets) + 2

        row = self._section(sheet, row, "担任授業期間", "開始日・終了日は元々期間を意味するルールです。条件表と組み合わせて対象クラスを決めます。")
        homeroom_rows = [[r.rule_id, r.rule_name, self._yn(r.enabled), r.start_date, r.end_date, ""] for r in data.homeroom_boundary_rules]
        row = self._add_table(sheet, row, "T_HOMEROOM_RULES", ["内部ID", "ルール", "使用", "開始日", "終了日", "備考"], homeroom_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"}) + 2
        condition_rows: list[list[Any]] = []
        field_display = {"class_id": "クラス", "division": "学部", "grade": "学年", "exam_category": "受験区分", "campus_id": "校舎", "has_regular_homeroom_lesson": "担任通常授業あり"}
        operator_display = {"eq": "=", "ne": "≠", "in": "いずれか", "ge": "以上", "le": "以下", "between": "範囲"}
        for rule in data.homeroom_boundary_rules:
            for field, operator, value in zip(rule.condition_fields, rule.condition_operators, rule.condition_values, strict=True):
                condition_rows.append([rule.rule_name, field_display.get(field, field), field, operator_display.get(operator, operator), operator, value])
        row = self._add_table(sheet, row, "T_HOMEROOM_RULE_CONDITIONS", ["ルール", "条件項目", "内部条件項目", "比較", "内部比較", "値"], condition_rows, hidden_columns={"内部条件項目", "内部比較"}) + 2

        row = self._section(sheet, row, "クラス組重複禁止", "ペアは明示指定です。クラス名から自動推定しません。")
        pair_rows = []
        for rule in data.class_pair_overlap_rules:
            campus_a, class_a = class_lookup[rule.first_class_id]
            campus_b, class_b = class_lookup[rule.second_class_id]
            pair_rows.append([rule.rule_id, rule.rule_name, self._yn(rule.enabled), campus_a, class_a, campus_b, class_b, ""])
        self._add_table(sheet, row, "T_CLASS_PAIRS", ["内部ID", "ルール", "使用", "校舎A", "クラスA", "校舎B", "クラスB", "備考"], pair_rows, yes_no_columns={"使用"}, hidden_columns={"内部ID"})
        self._finalize_sheet(sheet)

    def _write_check_sheet(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("08_入力確認")
        sheet["A1"] = "入力確認"
        sheet["A1"].font = Font(size=18, bold=True)
        sheet["A3"] = "このシートは確認用です。入力の正本ではありません。実行時Validationのエラーがある場合はその指示を優先してください。"
        summary = [
            ("校舎", len(data.campuses)),
            ("教室", len(data.rooms)),
            ("教師", len(data.teachers)),
            ("クラス", len(data.classes)),
            ("教科", len(data.subjects)),
            ("授業要求", len(data.lesson_requirements)),
            ("教師個別休み", len(data.teacher_leaves)),
            ("基本配置ルール", len(data.placement_rules)),
            ("Hard配置数segment", len(data.lesson_count_rule_segments)),
            ("Soft配置数segment", len(data.lesson_count_preference_rule_segments)),
            ("教師休日日数ルール", len(data.teacher_day_off_rules)),
            ("担任期間ルール", len(data.homeroom_boundary_rules)),
            ("クラス組", len(data.class_pair_overlap_rules)),
        ]
        self._add_table(sheet, 5, "T_INPUT_SUMMARY", ["確認項目", "件数"], summary)
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 18

    def _write_ai_guide(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("90_AI編集ガイド")
        sheet["A1"] = "AI編集ガイド / 入力契約 v2.0"
        sheet["A1"].font = Font(size=18, bold=True)
        guide = [
            ("最重要", "このExcelだけを渡されたAIでも安全に編集できることを目的とする。利用者の指示にない一般化・推論を行わない。"),
            ("○/×", "○=True/有効/対象、×=False/無効/対象外。Optional Booleanでは『指定なし』が第三の状態。"),
            (BLANK_TOKEN, f"意図的な空欄値。実セルの空欄とは異なる。{BLANK_TOKEN} を空セルへ置換しない。"),
            (UNSPECIFIED, "その項目では条件を設定しない。×とは異なる。"),
            ("日付集合", "T_DATE_SETS に明示された日付だけが集合員。最小日～最大日の全日を含むとは限らない。"),
            ("固定対象", "Hard/Softの対象はTARGETS表に明示されたクラス×教科だけ。学年・学部条件へ勝手に一般化しない。"),
            ("クラス識別", "校舎+クラス名で識別する。同名クラスが別校舎に存在し得る。"),
            ("教室識別", "校舎+教室名で識別する。"),
            ("内部列", "『内部』で始まる列はシステム保持用。通常は変更しない。"),
            ("例: 高宮6通を8/7不可", "開講カレンダー全体を休みにせず、対象クラス・教科に対する個別配置禁止ルールを変更する。"),
            ("例: 教師の1限6限休み", "教師個別休みで1限=○、2～5限=×、6限=○。1日休みへ一般化しない。"),
            (f"例: 担任を{BLANK_TOKEN}", f"担任セルで {BLANK_TOKEN} を明示選択する。セル自体を空にしない。"),
            ("例: 公休候補から1日除外", "T_DATE_SETS の該当する1行だけを削除する。期間の開始日・終了日は変更しない。"),
        ]
        self._add_table(sheet, 3, "T_AI_GUIDE", ["項目", "説明"], guide)
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 110
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _write_system(self, workbook: Workbook, data: InputDataModel) -> None:
        sheet = workbook.create_sheet("_system")
        rows = [["schema_version", "2.0"], ["source_schema_version", data.settings.schema_version], ["blank_token", BLANK_TOKEN], ["unspecified_token", UNSPECIFIED]]
        self._add_table(sheet, 1, "T_SYSTEM", ["項目", "値"], rows)

    def _group_lesson_count_rules(self, segments: Iterable[Any], class_lookup: dict[str, tuple[str, str]], subject_names: dict[str, str], periods: list[Any], *, soft: bool) -> tuple[list[list[Any]], list[list[Any]]]:
        by_rule: dict[str, list[Any]] = defaultdict(list)
        for segment in segments:
            by_rule[segment.rule_id].append(segment)
        signature_groups: dict[tuple[Any, ...], list[tuple[str, str]]] = defaultdict(list)
        segment_rows_by_signature: dict[tuple[Any, ...], list[Any]] = {}
        for rule_segments in by_rule.values():
            ordered = sorted(rule_segments, key=lambda item: item.segment_id)
            first = ordered[0]
            count = first.preferred_periods if soft else first.exact_periods
            signature = (
                first.rule_name,
                first.enabled,
                count,
                tuple((item.start_date, item.end_date, tuple(item.target_period_ids)) for item in ordered),
            )
            signature_groups[signature].append((first.class_id, first.subject_id))
            segment_rows_by_signature[signature] = ordered
        rule_rows: list[list[Any]] = []
        targets: list[list[Any]] = []
        for group_index, (signature, target_pairs) in enumerate(signature_groups.items(), start=1):
            rule_name, enabled, count, _ = signature
            display_name = f"{rule_name} [{group_index:02d}]" if sum(1 for key in signature_groups if key[0] == rule_name) > 1 else rule_name
            for segment_index, segment in enumerate(segment_rows_by_signature[signature], start=1):
                target_set = set(segment.target_period_ids)
                rule_rows.append([display_name, segment_index, self._yn(enabled), count, segment.start_date, segment.end_date, *[self._yn(period.period_id in target_set) for period in periods], ""])
            for class_id, subject_id in target_pairs:
                campus, class_name = class_lookup[class_id]
                targets.append([display_name, campus, class_name, subject_names[subject_id]])
        return rule_rows, targets

    @staticmethod
    def _date_sets(data: InputDataModel) -> dict[str, tuple[date, ...]]:
        result: dict[str, tuple[date, ...]] = {}
        reverse: dict[tuple[date, ...], str] = {}
        for rule in data.teacher_day_off_rules:
            dates = tuple(rule.eligible_dates)
            if dates not in reverse:
                name = f"休み候補_{len(reverse) + 1:02d}"
                reverse[dates] = name
                result[name] = dates
        return result

    def _section(self, sheet: Any, row: int, title: str, description: str) -> int:
        sheet.cell(row, 1, title)
        sheet.cell(row, 1).font = Font(size=14, bold=True)
        sheet.cell(row, 1).fill = self.TITLE_FILL
        sheet.cell(row + 1, 1, description)
        sheet.cell(row + 1, 1).alignment = Alignment(wrap_text=True)
        return row + 3

    def _add_table(self, sheet: Any, start_row: int, name: str, headers: list[str], rows: Iterable[Iterable[Any]], *, yes_no_columns: set[str] | None = None, optional_yes_no_columns: set[str] | None = None, hidden_columns: set[str] | None = None) -> int:
        yes_no_columns = yes_no_columns or set()
        optional_yes_no_columns = optional_yes_no_columns or set()
        hidden_columns = hidden_columns or set()
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(start_row, column, header)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        materialized = list(rows)
        if not materialized:
            materialized = [[None for _ in headers]]
        for offset, values in enumerate(materialized, start=1):
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(start_row + offset, column, value)
                header = headers[column - 1]
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header in hidden_columns:
                    cell.fill = self.AUTO_FILL
                    cell.protection = Protection(locked=True)
                else:
                    cell.fill = self.INPUT_FILL
        end_row = start_row + len(materialized)
        end_col = len(headers)
        reference = f"A{start_row}:{self._column_letter(end_col)}{end_row}"
        table = Table(displayName=name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
        for header in yes_no_columns:
            if header in headers:
                self._add_list_validation(sheet, headers.index(header) + 1, start_row + 1, end_row + 200, f'"{YES},{NO}"')
        for header in optional_yes_no_columns:
            if header in headers:
                self._add_list_validation(sheet, headers.index(header) + 1, start_row + 1, end_row + 200, f'"{YES},{NO},{UNSPECIFIED}"')
        for header in hidden_columns:
            if header in headers:
                sheet.column_dimensions[self._column_letter(headers.index(header) + 1)].hidden = True
        return end_row

    @staticmethod
    def _add_list_validation(sheet: Any, column: int, first_row: int, last_row: int, formula: str) -> None:
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "一覧から値を選択してください。"
        validation.errorTitle = "入力値が不正です"
        validation.prompt = "プルダウンから選択してください。"
        validation.promptTitle = "入力方法"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{ExcelV2WorkbookWriterAdapter._column_letter(column)}{first_row}:{ExcelV2WorkbookWriterAdapter._column_letter(column)}{last_row}")

    def _finalize_sheet(self, sheet: Any) -> None:
        sheet.freeze_panes = "A4"
        sheet.sheet_view.showGridLines = False
        for column in range(1, sheet.max_column + 1):
            letter = self._column_letter(column)
            if sheet.column_dimensions[letter].hidden:
                continue
            max_length = 10
            for cell in sheet[letter]:
                if cell.value is not None:
                    max_length = max(max_length, min(len(str(cell.value)) + 2, 32))
            sheet.column_dimensions[letter].width = max_length
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _yn(value: bool) -> str:
        return YES if value else NO

    @staticmethod
    def _display_optional(value: Any) -> Any:
        return UNSPECIFIED if value is None else value

    @staticmethod
    def _display_optional_bool(value: bool | None) -> str:
        if value is None:
            return UNSPECIFIED
        return YES if value else NO

    @staticmethod
    def _column_letter(index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
