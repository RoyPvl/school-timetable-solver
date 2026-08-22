from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from school_timetable_solver.adapter.excel_v2_workbook_postprocessor import (
    ExcelV2WorkbookPostprocessor,
)
from school_timetable_solver.model.input_models import InputDataModel


class ReferenceLabelExcelV2WorkbookPostprocessor(ExcelV2WorkbookPostprocessor):
    """Reference-label postprocessor for teachers, subjects, and classes.

    ``ExcelV2WorkbookWriterAdapter`` writes a single all-empty data row when a
    logical table has zero rows so that an Excel Table object can still exist.
    Reference rewriting ignores that physical placeholder row.
    """

    _exam_category_display = {
        "exam": "受験",
        "non_exam": "非受験",
        "special": "特別",
        "none": "なし",
    }

    def execute(self, path: Path, data: InputDataModel) -> None:
        # Parent handles teacher/subject labels, reference map creation, direct
        # class-campus completion, and general layout fixes.
        super().execute(path, data)

        workbook = load_workbook(path)
        try:
            class_labels = self._build_class_labels(data)
            self._rewrite_class_references(workbook, data, class_labels)
            self._append_class_reference_rows(workbook, data, class_labels)
            self._append_class_guidance(workbook)
            workbook.save(path)
        finally:
            workbook.close()

    @staticmethod
    def _nonblank_rows(
        worksheet: Any,
        headers: dict[str, int],
        min_row: int,
        max_row: int,
    ) -> list[int]:
        columns = tuple(headers.values())
        return [
            row
            for row in range(min_row + 1, max_row + 1)
            if any(worksheet.cell(row, column).value is not None for column in columns)
        ]

    def _rewrite_teacher_leaves(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "05_教師条件",
            "T_TEACHER_LEAVES",
        )
        excel_rows = self._nonblank_rows(worksheet, headers, min_row, max_row)
        if len(excel_rows) != len(data.teacher_leaves):
            raise ValueError("T_TEACHER_LEAVES row count does not match source model")
        for row, model in zip(excel_rows, data.teacher_leaves, strict=True):
            worksheet.cell(row, headers["教師"]).value = labels[model.teacher_id]

    def _rewrite_lesson_count_targets(
        self,
        workbook: Any,
        data: InputDataModel,
        subject_labels: dict[str, str],
        *,
        soft: bool,
    ) -> None:
        segments = (
            data.lesson_count_preference_rule_segments
            if soft
            else data.lesson_count_rule_segments
        )
        table_name = "T_LESSON_COUNT_SOFT_TARGETS" if soft else "T_LESSON_COUNT_HARD_TARGETS"
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "07_個別ルール",
            table_name,
        )
        expected_subject_ids = self._grouped_target_subject_ids(data, segments)
        excel_rows = self._nonblank_rows(worksheet, headers, min_row, max_row)
        if len(excel_rows) != len(expected_subject_ids):
            raise ValueError(f"{table_name} row count does not match source model")
        for row, subject_id in zip(excel_rows, expected_subject_ids, strict=True):
            worksheet.cell(row, headers["教科"]).value = subject_labels[subject_id]

    def _build_class_labels(self, data: InputDataModel) -> dict[str, str]:
        key_counts = Counter((item.campus_id, item.class_name) for item in data.classes)
        provisional: dict[str, str] = {}
        for item in data.classes:
            if key_counts[(item.campus_id, item.class_name)] == 1:
                provisional[item.class_id] = item.class_name
                continue
            context = self._exam_category_display.get(item.exam_category, item.exam_category)
            provisional[item.class_id] = f"{item.class_name}（{context}）"

        duplicate_counts = Counter(
            (item.campus_id, provisional[item.class_id])
            for item in data.classes
        )
        return {
            item.class_id: (
                provisional[item.class_id]
                if duplicate_counts[(item.campus_id, provisional[item.class_id])] == 1
                else f"{provisional[item.class_id]} [{item.class_id}]"
            )
            for item in data.classes
        }

    def _rewrite_class_references(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        self._rewrite_class_master(workbook, data, labels)
        self._rewrite_requirement_classes(workbook, data, labels)
        self._rewrite_placement_classes(workbook, data, labels)
        self._rewrite_lesson_count_class_targets(workbook, data, labels, soft=False)
        self._rewrite_lesson_count_class_targets(workbook, data, labels, soft=True)
        self._rewrite_class_pairs(workbook, data, labels)

    def _rewrite_class_master(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "04_クラス・授業",
            "T_CLASSES",
        )
        ids = {item.class_id for item in data.classes}
        for row in self._nonblank_rows(worksheet, headers, min_row, max_row):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            if internal_id in ids:
                worksheet.cell(row, headers["クラス"]).value = labels[internal_id]

    def _rewrite_requirement_classes(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "04_クラス・授業",
            "T_LESSON_REQUIREMENTS",
        )
        by_id = {item.requirement_id: item for item in data.lesson_requirements}
        for row in self._nonblank_rows(worksheet, headers, min_row, max_row):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = by_id.get(internal_id)
            if model is not None:
                worksheet.cell(row, headers["クラス"]).value = labels[model.class_id]

    def _rewrite_placement_classes(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "06_基本配置ルール",
            "T_PLACEMENT_RULES",
        )
        by_id = {item.rule_id: item for item in data.placement_rules}
        for row in self._nonblank_rows(worksheet, headers, min_row, max_row):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = by_id.get(internal_id)
            if model is None:
                continue
            direct_class_id = None
            for field, operator, value in zip(
                model.condition_fields,
                model.condition_operators,
                model.condition_values,
                strict=True,
            ):
                if field == "class_id" and operator == "eq" and value in labels:
                    direct_class_id = value
                    break
            if direct_class_id is not None:
                worksheet.cell(row, headers["クラス"]).value = labels[direct_class_id]

    def _rewrite_lesson_count_class_targets(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
        *,
        soft: bool,
    ) -> None:
        segments = (
            data.lesson_count_preference_rule_segments
            if soft
            else data.lesson_count_rule_segments
        )
        table_name = "T_LESSON_COUNT_SOFT_TARGETS" if soft else "T_LESSON_COUNT_HARD_TARGETS"
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "07_個別ルール",
            table_name,
        )
        expected_class_ids = self._grouped_target_class_ids(segments)
        excel_rows = self._nonblank_rows(worksheet, headers, min_row, max_row)
        if len(excel_rows) != len(expected_class_ids):
            raise ValueError(f"{table_name} class row count does not match source model")
        for row, class_id in zip(excel_rows, expected_class_ids, strict=True):
            worksheet.cell(row, headers["クラス"]).value = labels[class_id]

    def _rewrite_class_pairs(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(
            workbook,
            "07_個別ルール",
            "T_CLASS_PAIRS",
        )
        by_id = {item.rule_id: item for item in data.class_pair_overlap_rules}
        for row in self._nonblank_rows(worksheet, headers, min_row, max_row):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = by_id.get(internal_id)
            if model is None:
                continue
            worksheet.cell(row, headers["クラスA"]).value = labels[model.first_class_id]
            worksheet.cell(row, headers["クラスB"]).value = labels[model.second_class_id]

    def _grouped_target_class_ids(self, segments: Iterable[Any]) -> list[str]:
        by_rule: dict[str, list[Any]] = defaultdict(list)
        for segment in segments:
            by_rule[segment.rule_id].append(segment)
        signature_groups: dict[tuple[Any, ...], list[tuple[str, str]]] = defaultdict(list)
        for rule_segments in by_rule.values():
            ordered = sorted(rule_segments, key=lambda item: item.segment_id)
            first = ordered[0]
            count = getattr(first, "preferred_periods", None)
            if count is None:
                count = first.exact_periods
            signature = (
                first.rule_name,
                first.enabled,
                count,
                tuple(
                    (item.start_date, item.end_date, tuple(item.target_period_ids))
                    for item in ordered
                ),
            )
            signature_groups[signature].append((first.class_id, first.subject_id))
        return [class_id for pairs in signature_groups.values() for class_id, _ in pairs]

    def _append_class_reference_rows(
        self,
        workbook: Any,
        data: InputDataModel,
        labels: dict[str, str],
    ) -> None:
        worksheet = workbook["_system"]
        table = worksheet.tables["T_REFERENCE_MAP"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        start_row = max_row + 1
        for offset, item in enumerate(data.classes):
            row = start_row + offset
            worksheet.cell(row, min_col, "class")
            worksheet.cell(row, min_col + 1, item.class_id)
            worksheet.cell(row, min_col + 2, labels[item.class_id])
            worksheet.cell(row, min_col + 3, item.class_name)
        table.ref = f"A{min_row}:D{start_row + len(data.classes) - 1}"

    @staticmethod
    def _append_class_guidance(workbook: Any) -> None:
        intro = workbook["00_最初に読む"]
        row = intro.max_row + 1
        intro.cell(
            row,
            1,
            "・同一校舎に同じ出力名のクラスがある場合も、入力用ラベルで区別します。時間割出力では元のクラス名に戻ります。",
        )

        guide = workbook["90_AI編集ガイド"]
        row = guide.max_row + 1
        guide.cell(row, 1, "クラス参照")
        guide.cell(
            row,
            2,
            "同一校舎で同名クラスが存在しても入力用ラベルで区別する。出力名へ逆変換する際は T_REFERENCE_MAP を使う。",
        )
