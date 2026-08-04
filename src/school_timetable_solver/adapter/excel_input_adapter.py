from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from school_timetable_solver.model.input_models import (
    CalendarDayModel,
    InputDataModel,
    InputWorkbookSettingsModel,
    LessonCountPreferenceRuleSegmentModel,
    LessonCountRuleSegmentModel,
    LessonRequirementModel,
    PlacementRuleModel,
    TeacherDayOffRuleModel,
    TeacherLeaveModel,
)
from school_timetable_solver.model.master_models import (
    CampusModel,
    ClassModel,
    PeriodModel,
    RoomModel,
    SubjectModel,
    TeacherModel,
)
from school_timetable_solver.model.result_models import (
    InputReadResultModel,
    ValidationIssueModel,
)


class ExcelInputReaderAdapter:
    """Read input-contract v0.6 workbooks into Excel-independent models."""

    _required_headers: ClassVar[dict[str, tuple[str, ...]]] = {
        "01_基本設定": ("setting_key", "setting_value", "description"),
        "02_開講カレンダー": (
            "date",
            "output_enabled",
            "period_1_enabled",
            "period_2_enabled",
            "period_3_enabled",
            "period_4_enabled",
            "period_5_enabled",
            "period_6_enabled",
            "note",
        ),
        "03_時限": ("period_id", "period_name", "output_order", "start_time", "end_time"),
        "04_校舎": ("campus_id", "campus_name", "output_order", "enabled", "note"),
        "05_教室": (
            "room_id",
            "room_name",
            "campus_id",
            "output_order",
            "enabled",
            "note",
        ),
        "06_教師": (
            "teacher_id",
            "teacher_name",
            "home_campus_id",
            "enabled",
            "note",
        ),
        "07_クラス": (
            "class_id",
            "class_name",
            "campus_id",
            "division",
            "grade",
            "exam_category",
            "homeroom_teacher_id",
            "enabled",
            "note",
        ),
        "08_教科": ("subject_id", "subject_name", "enabled", "note"),
        "09_授業要求": (
            "requirement_id",
            "class_id",
            "subject_id",
            "teacher_id",
            "required_periods",
            "max_periods_per_day",
            "enabled",
            "note",
        ),
        "10_教師休み": (
            "teacher_id",
            "date",
            "unavailable_periods",
            "note",
        ),
        "11_配置ルール": (
            "rule_id",
            "rule_name",
            "enabled",
            "constraint_type",
            "target_entity",
            "condition_field",
            "condition_operator",
            "condition_value",
            "campus_id",
            "start_date",
            "end_date",
            "weekdays",
            "allowed_periods",
            "daily_hard_limit",
            "consecutive_limit",
            "attendance_streak_limit",
            "priority",
            "note",
            "preferred_attendance_streak_limit",
        ),
        "12_選好設定": ("rule_id", "enabled", "weight", "note"),
        "13_授業配置数ルール": (
            "rule_id",
            "segment_id",
            "rule_name",
            "enabled",
            "class_id",
            "subject_id",
            "exact_periods",
            "start_date",
            "end_date",
            "target_periods",
            "note",
        ),
        "14_授業配置数選好ルール": (
            "rule_id",
            "segment_id",
            "rule_name",
            "enabled",
            "class_id",
            "subject_id",
            "preferred_periods",
            "start_date",
            "end_date",
            "target_periods",
            "note",
        ),
        "15_教師休日日数ルール": (
            "rule_id",
            "teacher_id",
            "enabled",
            "start_date",
            "end_date",
            "required_days_off",
            "note",
        ),
    }
    _required_sheets: ClassVar[tuple[str, ...]] = (
        "00_操作説明",
        *_required_headers.keys(),
    )

    def read(self, path: Path) -> InputReadResultModel:
        issues: list[ValidationIssueModel] = []
        if not path.is_file():
            return InputReadResultModel(
                None,
                (self._issue("INPUT_FILE_NOT_FOUND", str(path), "ファイルが存在しません"),),
            )
        if path.suffix.lower() != ".xlsx":
            return InputReadResultModel(
                None,
                (self._issue("INPUT_FILE_TYPE", str(path), ".xlsxファイルを指定してください"),),
            )
        try:
            workbook = load_workbook(path, data_only=True)
        except (InvalidFileException, OSError, ValueError, KeyError) as exc:
            return InputReadResultModel(
                None,
                (self._issue("INPUT_WORKBOOK_ERROR", str(path), f"Workbookを開けません: {exc}"),),
            )

        for sheet_name in self._required_sheets:
            if sheet_name not in workbook.sheetnames:
                issues.append(
                    self._issue("REQUIRED_SHEET_MISSING", sheet_name, "必須シートが存在しません")
                )
        if self._has_errors(issues):
            return InputReadResultModel(None, tuple(issues))

        rows_by_sheet = {
            sheet_name: self._read_rows(
                workbook[sheet_name],
                headers,
                issues,
            )
            for sheet_name, headers in self._required_headers.items()
        }
        if self._has_errors(issues):
            return InputReadResultModel(None, tuple(issues))

        settings = self._read_settings(rows_by_sheet["01_基本設定"], issues)
        periods = self._read_periods(rows_by_sheet["03_時限"], issues)
        calendar_days = self._read_calendar(rows_by_sheet["02_開講カレンダー"], periods, issues)
        campuses = self._read_campuses(rows_by_sheet["04_校舎"], issues)
        rooms = self._read_rooms(rows_by_sheet["05_教室"], issues)
        teachers = self._read_teachers(rows_by_sheet["06_教師"], issues)
        classes = self._read_classes(rows_by_sheet["07_クラス"], issues)
        subjects = self._read_subjects(rows_by_sheet["08_教科"], issues)
        requirements = self._read_requirements(rows_by_sheet["09_授業要求"], issues)
        teacher_leaves = self._read_teacher_leaves(
            rows_by_sheet["10_教師休み"],
            periods,
            issues,
        )
        placement_rules = self._read_placement_rules(rows_by_sheet["11_配置ルール"], issues)
        lesson_count_rule_segments = self._read_lesson_count_rule_segments(
            rows_by_sheet["13_授業配置数ルール"],
            issues,
        )
        lesson_count_preference_rule_segments = self._read_lesson_count_preference_rule_segments(
            rows_by_sheet["14_授業配置数選好ルール"],
            issues,
        )
        teacher_day_off_rules = self._read_teacher_day_off_rules(
            rows_by_sheet["15_教師休日日数ルール"],
            issues,
        )
        if settings is None or self._has_errors(issues):
            return InputReadResultModel(None, tuple(issues))
        return InputReadResultModel(
            InputDataModel(
                settings=settings,
                calendar_days=tuple(calendar_days),
                periods=tuple(periods),
                campuses=tuple(campuses),
                rooms=tuple(rooms),
                teachers=tuple(teachers),
                classes=tuple(classes),
                subjects=tuple(subjects),
                lesson_requirements=tuple(requirements),
                teacher_leaves=tuple(teacher_leaves),
                placement_rules=tuple(placement_rules),
                lesson_count_rule_segments=tuple(lesson_count_rule_segments),
                lesson_count_preference_rule_segments=tuple(lesson_count_preference_rule_segments),
                teacher_day_off_rules=tuple(teacher_day_off_rules),
            ),
            tuple(issues),
        )

    def _read_rows(
        self,
        worksheet: Worksheet,
        required_headers: tuple[str, ...],
        issues: list[ValidationIssueModel],
    ) -> list[tuple[int, dict[str, object]]]:
        headers = [
            cell.value.strip() if isinstance(cell.value, str) else "" for cell in worksheet[1]
        ]
        header_index = {header: index for index, header in enumerate(headers) if header}
        for header in required_headers:
            if header not in header_index:
                issues.append(
                    self._issue(
                        "REQUIRED_COLUMN_MISSING",
                        worksheet.title,
                        f"必須列が存在しません: {header}",
                        f"{worksheet.title}!1",
                    )
                )
        if any(header not in header_index for header in required_headers):
            return []
        rows: list[tuple[int, dict[str, object]]] = []
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if all(cell.value is None for cell in cells):
                continue
            rows.append(
                (
                    row_number,
                    {
                        header: cells[index].value if index < len(cells) else None
                        for header, index in header_index.items()
                    },
                )
            )
        return rows

    def _read_settings(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> InputWorkbookSettingsModel | None:
        values: dict[str, tuple[int, object]] = {}
        known_keys = {"schema_version", "timetable_name", "description"}
        for row_number, row in rows:
            key = self._text(row["setting_key"], "01_基本設定", row_number, "setting_key", issues)
            if key is None:
                continue
            if key in values:
                issues.append(
                    self._cell_issue(
                        "DUPLICATE_SETTING_KEY",
                        "01_基本設定",
                        row_number,
                        "setting_key",
                        f"setting_keyが重複しています: {key}",
                    )
                )
            elif key not in known_keys:
                issues.append(
                    ValidationIssueModel(
                        "UNKNOWN_SETTING_KEY",
                        "WARNING",
                        key,
                        "未知の設定キーを読み飛ばしました",
                        f"01_基本設定!setting_key@{row_number}",
                    )
                )
            else:
                values[key] = (row_number, row["setting_value"])
        for key in ("schema_version", "timetable_name"):
            if key not in values:
                issues.append(
                    self._issue("REQUIRED_SETTING_MISSING", key, "必須設定が存在しません")
                )
        if "schema_version" not in values or "timetable_name" not in values:
            return None
        schema_version = self._text(
            values["schema_version"][1],
            "01_基本設定",
            values["schema_version"][0],
            "setting_value",
            issues,
        )
        timetable_name = self._text(
            values["timetable_name"][1],
            "01_基本設定",
            values["timetable_name"][0],
            "setting_value",
            issues,
        )
        description = (
            self._optional_text(values["description"][1]) if "description" in values else None
        )
        if schema_version is not None and schema_version != "0.6":
            issues.append(
                self._issue(
                    "UNSUPPORTED_SCHEMA_VERSION",
                    "schema_version",
                    f"対応外のschema_versionです: {schema_version}",
                )
            )
        if schema_version is None or timetable_name is None:
            return None
        return InputWorkbookSettingsModel(schema_version, timetable_name, description)

    def _read_periods(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[PeriodModel]:
        result: list[PeriodModel] = []
        for row_number, row in rows:
            period_id = self._text(row["period_id"], "03_時限", row_number, "period_id", issues)
            period_name = self._text(
                row["period_name"], "03_時限", row_number, "period_name", issues
            )
            output_order = self._integer(
                row["output_order"], "03_時限", row_number, "output_order", issues
            )
            start_time = self._time(row["start_time"], "03_時限", row_number, "start_time", issues)
            end_time = self._time(row["end_time"], "03_時限", row_number, "end_time", issues)
            if None not in (period_id, period_name, output_order, start_time, end_time):
                assert period_id is not None
                assert period_name is not None
                assert output_order is not None
                assert start_time is not None
                assert end_time is not None
                result.append(
                    PeriodModel(
                        period_id=period_id,
                        period_name=period_name,
                        output_order=output_order,
                        start_time=start_time,
                        end_time=end_time,
                    )
                )
        return result

    def _read_lesson_count_rule_segments(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[LessonCountRuleSegmentModel]:
        result: list[LessonCountRuleSegmentModel] = []
        for row_number, row in rows:
            target_periods = self._text(
                row["target_periods"],
                "13_授業配置数ルール",
                row_number,
                "target_periods",
                issues,
            )
            required = (
                self._text(
                    row["rule_id"],
                    "13_授業配置数ルール",
                    row_number,
                    "rule_id",
                    issues,
                ),
                self._text(
                    row["segment_id"],
                    "13_授業配置数ルール",
                    row_number,
                    "segment_id",
                    issues,
                ),
                self._text(
                    row["rule_name"],
                    "13_授業配置数ルール",
                    row_number,
                    "rule_name",
                    issues,
                ),
                self._boolean(
                    row["enabled"],
                    "13_授業配置数ルール",
                    row_number,
                    "enabled",
                    issues,
                ),
                self._text(
                    row["class_id"],
                    "13_授業配置数ルール",
                    row_number,
                    "class_id",
                    issues,
                ),
                self._text(
                    row["subject_id"],
                    "13_授業配置数ルール",
                    row_number,
                    "subject_id",
                    issues,
                ),
                self._integer(
                    row["exact_periods"],
                    "13_授業配置数ルール",
                    row_number,
                    "exact_periods",
                    issues,
                ),
                self._date(
                    row["start_date"],
                    "13_授業配置数ルール",
                    row_number,
                    "start_date",
                    issues,
                ),
                self._date(
                    row["end_date"],
                    "13_授業配置数ルール",
                    row_number,
                    "end_date",
                    issues,
                ),
                target_periods,
            )
            if None in required:
                continue
            (
                rule_id,
                segment_id,
                rule_name,
                enabled,
                class_id,
                subject_id,
                exact_periods,
                start_date,
                end_date,
                target_periods,
            ) = required
            assert isinstance(rule_id, str)
            assert isinstance(segment_id, str)
            assert isinstance(rule_name, str)
            assert isinstance(enabled, bool)
            assert isinstance(class_id, str)
            assert isinstance(subject_id, str)
            assert isinstance(exact_periods, int)
            assert isinstance(start_date, date)
            assert isinstance(end_date, date)
            assert isinstance(target_periods, str)
            result.append(
                LessonCountRuleSegmentModel(
                    rule_id=rule_id,
                    segment_id=segment_id,
                    rule_name=rule_name,
                    enabled=enabled,
                    class_id=class_id,
                    subject_id=subject_id,
                    exact_periods=exact_periods,
                    start_date=start_date,
                    end_date=end_date,
                    target_period_ids=tuple(
                        part.strip() for part in target_periods.split("|") if part.strip()
                    ),
                )
            )
        return result

    def _read_lesson_count_preference_rule_segments(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[LessonCountPreferenceRuleSegmentModel]:
        sheet_name = "14_授業配置数選好ルール"
        result: list[LessonCountPreferenceRuleSegmentModel] = []
        for row_number, row in rows:
            target_periods = self._text(
                row["target_periods"],
                sheet_name,
                row_number,
                "target_periods",
                issues,
            )
            required = (
                self._text(row["rule_id"], sheet_name, row_number, "rule_id", issues),
                self._text(row["segment_id"], sheet_name, row_number, "segment_id", issues),
                self._text(row["rule_name"], sheet_name, row_number, "rule_name", issues),
                self._boolean(row["enabled"], sheet_name, row_number, "enabled", issues),
                self._text(row["class_id"], sheet_name, row_number, "class_id", issues),
                self._text(row["subject_id"], sheet_name, row_number, "subject_id", issues),
                self._integer(
                    row["preferred_periods"],
                    sheet_name,
                    row_number,
                    "preferred_periods",
                    issues,
                ),
                self._date(row["start_date"], sheet_name, row_number, "start_date", issues),
                self._date(row["end_date"], sheet_name, row_number, "end_date", issues),
                target_periods,
            )
            if None in required:
                continue
            (
                rule_id,
                segment_id,
                rule_name,
                enabled,
                class_id,
                subject_id,
                preferred_periods,
                start_date,
                end_date,
                target_periods,
            ) = required
            assert isinstance(rule_id, str)
            assert isinstance(segment_id, str)
            assert isinstance(rule_name, str)
            assert isinstance(enabled, bool)
            assert isinstance(class_id, str)
            assert isinstance(subject_id, str)
            assert isinstance(preferred_periods, int)
            assert isinstance(start_date, date)
            assert isinstance(end_date, date)
            assert isinstance(target_periods, str)
            result.append(
                LessonCountPreferenceRuleSegmentModel(
                    rule_id=rule_id,
                    segment_id=segment_id,
                    rule_name=rule_name,
                    enabled=enabled,
                    class_id=class_id,
                    subject_id=subject_id,
                    preferred_periods=preferred_periods,
                    start_date=start_date,
                    end_date=end_date,
                    target_period_ids=tuple(
                        part.strip() for part in target_periods.split("|") if part.strip()
                    ),
                )
            )
        return result

    def _read_calendar(
        self,
        rows: list[tuple[int, dict[str, object]]],
        periods: list[PeriodModel],
        issues: list[ValidationIssueModel],
    ) -> list[CalendarDayModel]:
        periods_by_order = {period.output_order: period.period_id for period in periods}
        result: list[CalendarDayModel] = []
        for row_number, row in rows:
            target_date = self._date(row["date"], "02_開講カレンダー", row_number, "date", issues)
            output_enabled = self._boolean(
                row["output_enabled"],
                "02_開講カレンダー",
                row_number,
                "output_enabled",
                issues,
            )
            flags = [
                self._boolean(
                    row[f"period_{order}_enabled"],
                    "02_開講カレンダー",
                    row_number,
                    f"period_{order}_enabled",
                    issues,
                )
                for order in range(1, 7)
            ]
            if target_date is not None and output_enabled is not None and None not in flags:
                result.append(
                    CalendarDayModel(
                        target_date=target_date,
                        output_enabled=output_enabled,
                        enabled_period_ids=tuple(
                            periods_by_order[order]
                            for order, enabled in enumerate(flags, start=1)
                            if enabled and order in periods_by_order
                        ),
                        note=self._optional_text(row["note"]),
                    )
                )
        return result

    def _read_campuses(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[CampusModel]:
        result: list[CampusModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["campus_id"], "04_校舎", row_number, "campus_id", issues),
                self._text(row["campus_name"], "04_校舎", row_number, "campus_name", issues),
                self._integer(row["output_order"], "04_校舎", row_number, "output_order", issues),
                self._boolean(row["enabled"], "04_校舎", row_number, "enabled", issues),
            )
            if None not in values:
                campus_id, campus_name, output_order, enabled = values
                assert campus_id is not None
                assert campus_name is not None
                assert output_order is not None
                assert enabled is not None
                result.append(CampusModel(campus_id, campus_name, output_order, enabled))
        return result

    def _read_rooms(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[RoomModel]:
        result: list[RoomModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["room_id"], "05_教室", row_number, "room_id", issues),
                self._text(row["room_name"], "05_教室", row_number, "room_name", issues),
                self._text(row["campus_id"], "05_教室", row_number, "campus_id", issues),
                self._integer(row["output_order"], "05_教室", row_number, "output_order", issues),
                self._boolean(row["enabled"], "05_教室", row_number, "enabled", issues),
            )
            if None not in values:
                room_id, room_name, campus_id, output_order, enabled = values
                assert room_id is not None
                assert room_name is not None
                assert campus_id is not None
                assert output_order is not None
                assert enabled is not None
                result.append(RoomModel(room_id, room_name, campus_id, output_order, enabled))
        return result

    def _read_teachers(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[TeacherModel]:
        result: list[TeacherModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["teacher_id"], "06_教師", row_number, "teacher_id", issues),
                self._teacher_name(row["teacher_name"], row_number, issues),
                self._text(
                    row["home_campus_id"],
                    "06_教師",
                    row_number,
                    "home_campus_id",
                    issues,
                ),
                self._boolean(row["enabled"], "06_教師", row_number, "enabled", issues),
            )
            if None not in values:
                teacher_id, teacher_name, home_campus_id, enabled = values
                assert teacher_id is not None
                assert teacher_name is not None
                assert home_campus_id is not None
                assert enabled is not None
                result.append(
                    TeacherModel(
                        teacher_id,
                        teacher_name,
                        home_campus_id,
                        enabled,
                    )
                )
        return result

    def _teacher_name(
        self,
        value: object,
        row: int,
        issues: list[ValidationIssueModel],
    ) -> str | None:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        issues.append(
            self._cell_issue(
                "REQUIRED_TEXT_INVALID",
                "06_教師",
                row,
                "teacher_name",
                "teacher_nameは文字列型で入力してください",
            )
        )
        return None

    def _subject_name(
        self,
        value: object,
        row: int,
        issues: list[ValidationIssueModel],
    ) -> str | None:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        issues.append(
            self._cell_issue(
                "REQUIRED_TEXT_INVALID",
                "08_教科",
                row,
                "subject_name",
                "subject_nameは文字列型で入力してください",
            )
        )
        return None

    def _read_classes(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[ClassModel]:
        result: list[ClassModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["class_id"], "07_クラス", row_number, "class_id", issues),
                self._text(row["class_name"], "07_クラス", row_number, "class_name", issues),
                self._text(row["campus_id"], "07_クラス", row_number, "campus_id", issues),
                self._text(row["division"], "07_クラス", row_number, "division", issues),
                self._integer(row["grade"], "07_クラス", row_number, "grade", issues),
                self._text(row["exam_category"], "07_クラス", row_number, "exam_category", issues),
                self._optional_text(row["homeroom_teacher_id"]),
                self._boolean(row["enabled"], "07_クラス", row_number, "enabled", issues),
            )
            if all(value is not None for value in (*values[:6], values[7])):
                (
                    class_id,
                    class_name,
                    campus_id,
                    division,
                    grade,
                    exam_category,
                    homeroom_teacher_id,
                    enabled,
                ) = values
                assert class_id is not None
                assert class_name is not None
                assert campus_id is not None
                assert division is not None
                assert grade is not None
                assert exam_category is not None
                assert enabled is not None
                result.append(
                    ClassModel(
                        class_id,
                        class_name,
                        campus_id,
                        division,
                        grade,
                        exam_category,
                        homeroom_teacher_id,
                        enabled,
                    )
                )
        return result

    def _read_subjects(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[SubjectModel]:
        result: list[SubjectModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["subject_id"], "08_教科", row_number, "subject_id", issues),
                self._subject_name(row["subject_name"], row_number, issues),
                self._boolean(row["enabled"], "08_教科", row_number, "enabled", issues),
            )
            if None not in values:
                subject_id, subject_name, enabled = values
                assert subject_id is not None
                assert subject_name is not None
                assert enabled is not None
                result.append(SubjectModel(subject_id, subject_name, enabled))
        return result

    def _read_requirements(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[LessonRequirementModel]:
        result: list[LessonRequirementModel] = []
        for row_number, row in rows:
            values = (
                self._text(
                    row["requirement_id"], "09_授業要求", row_number, "requirement_id", issues
                ),
                self._text(row["class_id"], "09_授業要求", row_number, "class_id", issues),
                self._text(row["subject_id"], "09_授業要求", row_number, "subject_id", issues),
                self._text(row["teacher_id"], "09_授業要求", row_number, "teacher_id", issues),
                self._integer(
                    row["required_periods"],
                    "09_授業要求",
                    row_number,
                    "required_periods",
                    issues,
                ),
                self._optional_integer(
                    row["max_periods_per_day"],
                    "09_授業要求",
                    row_number,
                    "max_periods_per_day",
                    issues,
                ),
                self._boolean(row["enabled"], "09_授業要求", row_number, "enabled", issues),
            )
            if all(value is not None for value in (*values[:5], values[6])):
                (
                    requirement_id,
                    class_id,
                    subject_id,
                    teacher_id,
                    required_periods,
                    max_periods_per_day,
                    enabled,
                ) = values
                assert requirement_id is not None
                assert class_id is not None
                assert subject_id is not None
                assert teacher_id is not None
                assert required_periods is not None
                assert enabled is not None
                result.append(
                    LessonRequirementModel(
                        requirement_id,
                        class_id,
                        subject_id,
                        teacher_id,
                        required_periods,
                        max_periods_per_day,
                        enabled,
                    )
                )
        return result

    def _read_teacher_leaves(
        self,
        rows: list[tuple[int, dict[str, object]]],
        periods: list[PeriodModel],
        issues: list[ValidationIssueModel],
    ) -> list[TeacherLeaveModel]:
        period_ids = tuple(period.period_id for period in periods)
        result: list[TeacherLeaveModel] = []
        for row_number, row in rows:
            teacher_id = self._text(
                row["teacher_id"], "10_教師休み", row_number, "teacher_id", issues
            )
            target_date = self._date(row["date"], "10_教師休み", row_number, "date", issues)
            unavailable_periods = self._text(
                row["unavailable_periods"],
                "10_教師休み",
                row_number,
                "unavailable_periods",
                issues,
            )
            unavailable_period_ids: tuple[str, ...] = ()
            if unavailable_periods is not None:
                values = tuple(
                    part.strip() for part in unavailable_periods.split("|") if part.strip()
                )
                if "ALL" in values:
                    if values == ("ALL",):
                        unavailable_period_ids = period_ids
                    else:
                        issues.append(
                            self._cell_issue(
                                "INVALID_TEACHER_LEAVE_PERIODS",
                                "10_教師休み",
                                row_number,
                                "unavailable_periods",
                                "ALLは個別の時限IDと併記できません",
                            )
                        )
                else:
                    unavailable_period_ids = values
            if (
                teacher_id is not None
                and target_date is not None
                and unavailable_periods is not None
                and unavailable_period_ids
            ):
                result.append(
                    TeacherLeaveModel(
                        teacher_id=teacher_id,
                        target_date=target_date,
                        unavailable_period_ids=unavailable_period_ids,
                    )
                )
        return result

    def _read_teacher_day_off_rules(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[TeacherDayOffRuleModel]:
        sheet_name = "15_教師休日日数ルール"
        result: list[TeacherDayOffRuleModel] = []
        for row_number, row in rows:
            values = (
                self._text(row["rule_id"], sheet_name, row_number, "rule_id", issues),
                self._text(row["teacher_id"], sheet_name, row_number, "teacher_id", issues),
                self._boolean(row["enabled"], sheet_name, row_number, "enabled", issues),
                self._date(row["start_date"], sheet_name, row_number, "start_date", issues),
                self._date(row["end_date"], sheet_name, row_number, "end_date", issues),
                self._integer(
                    row["required_days_off"],
                    sheet_name,
                    row_number,
                    "required_days_off",
                    issues,
                ),
            )
            if None in values:
                continue
            rule_id, teacher_id, enabled, start_date, end_date, required_days_off = values
            assert isinstance(rule_id, str)
            assert isinstance(teacher_id, str)
            assert isinstance(enabled, bool)
            assert isinstance(start_date, date)
            assert isinstance(end_date, date)
            assert isinstance(required_days_off, int)
            result.append(
                TeacherDayOffRuleModel(
                    rule_id=rule_id,
                    teacher_id=teacher_id,
                    enabled=enabled,
                    start_date=start_date,
                    end_date=end_date,
                    required_days_off=required_days_off,
                )
            )
        return result

    def _read_placement_rules(
        self,
        rows: list[tuple[int, dict[str, object]]],
        issues: list[ValidationIssueModel],
    ) -> list[PlacementRuleModel]:
        result: list[PlacementRuleModel] = []
        for row_number, row in rows:
            required = (
                self._text(row["rule_id"], "11_配置ルール", row_number, "rule_id", issues),
                self._text(row["rule_name"], "11_配置ルール", row_number, "rule_name", issues),
                self._boolean(row["enabled"], "11_配置ルール", row_number, "enabled", issues),
                self._text(
                    row["constraint_type"],
                    "11_配置ルール",
                    row_number,
                    "constraint_type",
                    issues,
                ),
                self._text(
                    row["target_entity"],
                    "11_配置ルール",
                    row_number,
                    "target_entity",
                    issues,
                ),
                self._integer(row["priority"], "11_配置ルール", row_number, "priority", issues),
            )
            if None in required:
                continue
            rule_id, rule_name, enabled, constraint_type, target_entity, priority = required
            assert rule_id is not None
            assert rule_name is not None
            assert enabled is not None
            assert constraint_type is not None
            assert target_entity is not None
            assert priority is not None
            result.append(
                PlacementRuleModel(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    enabled=enabled,
                    constraint_type=constraint_type,
                    target_entity=target_entity,
                    condition_fields=self._pipe(row["condition_field"]),
                    condition_operators=self._pipe(row["condition_operator"]),
                    condition_values=self._pipe(row["condition_value"]),
                    campus_id=self._optional_text(row["campus_id"]),
                    start_date=self._optional_date(
                        row["start_date"],
                        "11_配置ルール",
                        row_number,
                        "start_date",
                        issues,
                    ),
                    end_date=self._optional_date(
                        row["end_date"],
                        "11_配置ルール",
                        row_number,
                        "end_date",
                        issues,
                    ),
                    weekdays=self._pipe(row["weekdays"]),
                    allowed_period_ids=self._pipe(row["allowed_periods"]),
                    daily_hard_limit=self._optional_integer(
                        row["daily_hard_limit"],
                        "11_配置ルール",
                        row_number,
                        "daily_hard_limit",
                        issues,
                    ),
                    consecutive_limit=self._optional_integer(
                        row["consecutive_limit"],
                        "11_配置ルール",
                        row_number,
                        "consecutive_limit",
                        issues,
                    ),
                    attendance_streak_limit=self._optional_integer(
                        row["attendance_streak_limit"],
                        "11_配置ルール",
                        row_number,
                        "attendance_streak_limit",
                        issues,
                    ),
                    priority=priority,
                    preferred_attendance_streak_limit=self._optional_integer(
                        row["preferred_attendance_streak_limit"],
                        "11_配置ルール",
                        row_number,
                        "preferred_attendance_streak_limit",
                        issues,
                    ),
                )
            )
        return result

    def _text(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> str | None:
        if isinstance(value, str) and value.strip():
            return value.strip()
        issues.append(
            self._cell_issue(
                "REQUIRED_TEXT_INVALID",
                sheet,
                row,
                header,
                "必須文字列が空欄または文字列型ではありません",
            )
        )
        return None

    def _integer(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> int | None:
        parsed: int | None = None
        if isinstance(value, int) and not isinstance(value, bool):
            parsed = value
        elif isinstance(value, float) and value.is_integer():
            parsed = int(value)
        if parsed is None:
            issues.append(
                self._cell_issue(
                    "INVALID_INTEGER", sheet, row, header, f"整数型ではありません: {value}"
                )
            )
        return parsed

    def _optional_integer(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> int | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        return self._integer(value, sheet, row, header, issues)

    def _boolean(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> bool | None:
        if isinstance(value, bool):
            return value
        issues.append(
            self._cell_issue(
                "INVALID_BOOLEAN",
                sheet,
                row,
                header,
                "ExcelのBoolean値TRUEまたはFALSEだけを使用してください",
            )
        )
        return None

    def _date(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> date | None:
        if isinstance(value, datetime):
            if value.time() == time():
                return value.date()
        elif isinstance(value, date):
            return value
        issues.append(
            self._cell_issue(
                "INVALID_DATE",
                sheet,
                row,
                header,
                "時刻部分を含まないExcel日付型を使用してください",
            )
        )
        return None

    def _optional_date(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> date | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        return self._date(value, sheet, row, header, issues)

    def _time(
        self,
        value: object,
        sheet: str,
        row: int,
        header: str,
        issues: list[ValidationIssueModel],
    ) -> time | None:
        if isinstance(value, time):
            return value
        issues.append(
            self._cell_issue(
                "INVALID_TIME",
                sheet,
                row,
                header,
                "日付部分を含まないExcel時刻型を使用してください",
            )
        )
        return None

    def _optional_text(self, value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() or None
        return str(value).strip() or None

    def _pipe(self, value: object) -> tuple[str, ...]:
        text = self._optional_text(value)
        if text is None:
            return ()
        return tuple(part.strip() for part in text.split("|") if part.strip())

    def _cell_issue(
        self,
        rule_id: str,
        sheet: str,
        row: int,
        header: str,
        message: str,
    ) -> ValidationIssueModel:
        return self._issue(rule_id, f"{sheet}.{header}", message, f"{sheet}!{header}@{row}")

    def _issue(
        self,
        rule_id: str,
        target: str,
        message: str,
        cell: str | None = None,
    ) -> ValidationIssueModel:
        return ValidationIssueModel(rule_id, "ERROR", target, message, cell)

    def _has_errors(self, issues: list[ValidationIssueModel]) -> bool:
        return any(issue.severity == "ERROR" for issue in issues)
