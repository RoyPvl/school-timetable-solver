from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from school_timetable_solver.adapter.excel_input_v2_adapter import UNSPECIFIED
from school_timetable_solver.model.input_models import InputDataModel


class ExcelV2WorkbookPostprocessor:
    """Apply source-model-aware fixes to a migrated v2 workbook.

    The public v2 workbook uses human-facing *input labels* for references while
    the domain model keeps stable internal IDs and output display names. This
    postprocessor creates the explicit label -> ID -> output-name reference map
    and rewrites generated reference cells to those labels. No label text is
    interpreted by the reader.
    """

    AUTO_FILL = PatternFill("solid", fgColor="E7E6E6")

    def execute(self, path: Path, data: InputDataModel) -> None:
        workbook = load_workbook(path)
        try:
            for worksheet in workbook.worksheets:
                if worksheet.title == "_system":
                    continue
                # Several tables are stacked vertically on the same worksheet.
                # Hiding a physical worksheet column for one table would also hide
                # unrelated fields in the other tables, so internal columns remain
                # visible and are visually marked instead.
                for dimension in worksheet.column_dimensions.values():
                    dimension.hidden = False

            self._fill_direct_class_rule_campuses(workbook, data)
            teacher_labels = self._build_teacher_labels(data)
            subject_labels = self._build_subject_labels(data)
            self._rewrite_reference_cells(workbook, data, teacher_labels, subject_labels)
            self._write_reference_map(workbook, data, teacher_labels, subject_labels)
            self._write_reference_guidance(workbook)
            workbook.save(path)
        finally:
            workbook.close()

    def _fill_direct_class_rule_campuses(self, workbook: Any, data: InputDataModel) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "06_基本配置ルール", "T_PLACEMENT_RULES")
        class_by_id = {item.class_id: item for item in data.classes}
        campus_name_by_id = {item.campus_id: item.campus_name for item in data.campuses}
        rule_by_id = {item.rule_id: item for item in data.placement_rules}

        for row in range(min_row + 1, max_row + 1):
            rule_id = worksheet.cell(row, headers["内部ID"]).value
            direct_class = worksheet.cell(row, headers["クラス"]).value
            campus_cell = worksheet.cell(row, headers["校舎"])
            if direct_class in (None, "", UNSPECIFIED) or campus_cell.value != UNSPECIFIED:
                continue
            rule = rule_by_id.get(str(rule_id))
            if rule is None:
                continue
            for field, operator, value in zip(
                rule.condition_fields,
                rule.condition_operators,
                rule.condition_values,
                strict=True,
            ):
                if field == "class_id" and operator == "eq" and value in class_by_id:
                    class_model = class_by_id[value]
                    campus_cell.value = campus_name_by_id[class_model.campus_id]
                    break

    def _build_teacher_labels(self, data: InputDataModel) -> dict[str, str]:
        campus_names = {item.campus_id: item.campus_name for item in data.campuses}
        class_by_id = {item.class_id: item for item in data.classes}
        contexts: dict[str, list[str]] = defaultdict(list)
        for requirement in data.lesson_requirements:
            class_model = class_by_id[requirement.class_id]
            context = f"{campus_names[class_model.campus_id]}/{class_model.class_name}"
            contexts[requirement.teacher_id].append(context)

        records: list[tuple[str, str, str, bool]] = []
        for teacher in data.teachers:
            output_name = teacher.teacher_name or ""
            base = output_name or "表示なし教師"
            candidates = sorted(dict.fromkeys(contexts.get(teacher.teacher_id, ())))
            context = candidates[0] if candidates else campus_names[teacher.home_campus_id]
            records.append((teacher.teacher_id, base, context, not bool(output_name)))
        return self._make_unique_labels(records)

    def _build_subject_labels(self, data: InputDataModel) -> dict[str, str]:
        campus_names = {item.campus_id: item.campus_name for item in data.campuses}
        class_by_id = {item.class_id: item for item in data.classes}
        contexts: dict[str, list[str]] = defaultdict(list)
        for requirement in data.lesson_requirements:
            class_model = class_by_id[requirement.class_id]
            context = f"{campus_names[class_model.campus_id]}/{class_model.class_name}"
            contexts[requirement.subject_id].append(context)

        records: list[tuple[str, str, str, bool]] = []
        for subject in data.subjects:
            output_name = subject.subject_name or ""
            base = output_name or "表示なし教科"
            candidates = sorted(dict.fromkeys(contexts.get(subject.subject_id, ())))
            context = candidates[0] if candidates else subject.lesson_type
            records.append((subject.subject_id, base, context, not bool(output_name)))
        return self._make_unique_labels(records)

    @staticmethod
    def _make_unique_labels(records: Iterable[tuple[str, str, str, bool]]) -> dict[str, str]:
        materialized = list(records)
        base_counts = Counter(base for _, base, _, _ in materialized)
        provisional: dict[str, str] = {}
        for internal_id, base, context, force_context in materialized:
            if force_context or base_counts[base] > 1:
                provisional[internal_id] = f"{base}（{context}）"
            else:
                provisional[internal_id] = base

        value_counts = Counter(provisional.values())
        result: dict[str, str] = {}
        for internal_id, label in provisional.items():
            result[internal_id] = label if value_counts[label] == 1 else f"{label} [{internal_id}]"
        return result

    def _rewrite_reference_cells(
        self,
        workbook: Any,
        data: InputDataModel,
        teacher_labels: dict[str, str],
        subject_labels: dict[str, str],
    ) -> None:
        self._rewrite_teacher_master(workbook, data, teacher_labels)
        self._rewrite_subject_master(workbook, data, subject_labels)
        self._rewrite_classes(workbook, data, teacher_labels)
        self._rewrite_requirements(workbook, data, teacher_labels, subject_labels)
        self._rewrite_teacher_leaves(workbook, data, teacher_labels)
        self._rewrite_day_off_rules(workbook, data, teacher_labels)
        self._rewrite_placement_rules(workbook, data, teacher_labels)
        self._rewrite_lesson_count_targets(workbook, data, subject_labels, soft=False)
        self._rewrite_lesson_count_targets(workbook, data, subject_labels, soft=True)

    def _rewrite_teacher_master(self, workbook: Any, data: InputDataModel, labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "03_教師", "T_TEACHERS")
        model_by_id = {item.teacher_id: item for item in data.teachers}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            if internal_id not in model_by_id:
                continue
            model = model_by_id[internal_id]
            worksheet.cell(row, headers["教師"]).value = labels[internal_id]
            worksheet.cell(row, headers["内部表示名"]).value = model.teacher_name or None
        header_cell = worksheet.cell(min_row, headers["内部表示名"])
        header_cell.value = "出力表示名"
        header_cell.fill = self.AUTO_FILL

    def _rewrite_subject_master(self, workbook: Any, data: InputDataModel, labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "04_クラス・授業", "T_SUBJECTS")
        model_by_id = {item.subject_id: item for item in data.subjects}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            if internal_id not in model_by_id:
                continue
            model = model_by_id[internal_id]
            worksheet.cell(row, headers["教科"]).value = labels[internal_id]
            worksheet.cell(row, headers["内部表示名"]).value = model.subject_name or None
        header_cell = worksheet.cell(min_row, headers["内部表示名"])
        header_cell.value = "出力表示名"
        header_cell.fill = self.AUTO_FILL

    def _rewrite_classes(self, workbook: Any, data: InputDataModel, teacher_labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "04_クラス・授業", "T_CLASSES")
        model_by_id = {item.class_id: item for item in data.classes}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = model_by_id.get(internal_id)
            if model is None:
                continue
            worksheet.cell(row, headers["担任"]).value = (
                teacher_labels[model.homeroom_teacher_id]
                if model.homeroom_teacher_id is not None
                else UNSPECIFIED
            )

    def _rewrite_requirements(
        self,
        workbook: Any,
        data: InputDataModel,
        teacher_labels: dict[str, str],
        subject_labels: dict[str, str],
    ) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "04_クラス・授業", "T_LESSON_REQUIREMENTS")
        model_by_id = {item.requirement_id: item for item in data.lesson_requirements}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = model_by_id.get(internal_id)
            if model is None:
                continue
            worksheet.cell(row, headers["教師"]).value = teacher_labels[model.teacher_id]
            worksheet.cell(row, headers["教科"]).value = subject_labels[model.subject_id]

    def _rewrite_teacher_leaves(self, workbook: Any, data: InputDataModel, labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "05_教師条件", "T_TEACHER_LEAVES")
        excel_rows = list(range(min_row + 1, max_row + 1))
        if len(excel_rows) != len(data.teacher_leaves):
            raise ValueError("T_TEACHER_LEAVES row count does not match source model")
        for row, model in zip(excel_rows, data.teacher_leaves, strict=True):
            worksheet.cell(row, headers["教師"]).value = labels[model.teacher_id]

    def _rewrite_day_off_rules(self, workbook: Any, data: InputDataModel, labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "05_教師条件", "T_TEACHER_DAY_OFF_RULES")
        model_by_id = {item.rule_id: item for item in data.teacher_day_off_rules}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = model_by_id.get(internal_id)
            if model is not None:
                worksheet.cell(row, headers["教師"]).value = labels[model.teacher_id]

    def _rewrite_placement_rules(self, workbook: Any, data: InputDataModel, labels: dict[str, str]) -> None:
        worksheet, headers, min_row, max_row = self._table(workbook, "06_基本配置ルール", "T_PLACEMENT_RULES")
        model_by_id = {item.rule_id: item for item in data.placement_rules}
        for row in range(min_row + 1, max_row + 1):
            internal_id = str(worksheet.cell(row, headers["内部ID"]).value)
            model = model_by_id.get(internal_id)
            if model is None:
                continue
            direct_teacher_id = None
            for field, operator, value in zip(
                model.condition_fields,
                model.condition_operators,
                model.condition_values,
                strict=True,
            ):
                if field == "teacher_id" and operator == "eq" and value in labels:
                    direct_teacher_id = value
                    break
            worksheet.cell(row, headers["教師"]).value = (
                labels[direct_teacher_id] if direct_teacher_id is not None else UNSPECIFIED
            )

    def _rewrite_lesson_count_targets(
        self,
        workbook: Any,
        data: InputDataModel,
        subject_labels: dict[str, str],
        *,
        soft: bool,
    ) -> None:
        segments = (
            data.lesson_count_preference_rule_segments
            if soft
            else data.lesson_count_rule_segments
        )
        table_name = "T_LESSON_COUNT_SOFT_TARGETS" if soft else "T_LESSON_COUNT_HARD_TARGETS"
        worksheet, headers, min_row, max_row = self._table(workbook, "07_個別ルール", table_name)
        expected_subject_ids = self._grouped_target_subject_ids(data, segments)
        excel_rows = list(range(min_row + 1, max_row + 1))
        if len(excel_rows) != len(expected_subject_ids):
            raise ValueError(f"{table_name} row count does not match source model")
        for row, subject_id in zip(excel_rows, expected_subject_ids, strict=True):
            worksheet.cell(row, headers["教科"]).value = subject_labels[subject_id]

    def _grouped_target_subject_ids(self, data: InputDataModel, segments: Iterable[Any]) -> list[str]:
        by_rule: dict[str, list[Any]] = defaultdict(list)
        for segment in segments:
            by_rule[segment.rule_id].append(segment)
        signature_groups: dict[tuple[Any, ...], list[tuple[str, str]]] = defaultdict(list)
        for rule_segments in by_rule.values():
            ordered = sorted(rule_segments, key=lambda item: item.segment_id)
            first = ordered[0]
            count = getattr(first, "preferred_periods", None)
            if count is None:
                count = first.exact_periods
            signature = (
                first.rule_name,
                first.enabled,
                count,
                tuple((item.start_date, item.end_date, tuple(item.target_period_ids)) for item in ordered),
            )
            signature_groups[signature].append((first.class_id, first.subject_id))
        return [subject_id for target_pairs in signature_groups.values() for _, subject_id in target_pairs]

    def _write_reference_map(
        self,
        workbook: Any,
        data: InputDataModel,
        teacher_labels: dict[str, str],
        subject_labels: dict[str, str],
    ) -> None:
        worksheet = workbook["_system"]
        if "T_REFERENCE_MAP" in worksheet.tables:
            del worksheet.tables["T_REFERENCE_MAP"]
        start_row = max(worksheet.max_row + 3, 8)
        headers = ["参照種別", "内部ID", "入力用ラベル", "出力表示名"]
        rows = [
            ["teacher", item.teacher_id, teacher_labels[item.teacher_id], item.teacher_name or None]
            for item in data.teachers
        ] + [
            ["subject", item.subject_id, subject_labels[item.subject_id], item.subject_name or None]
            for item in data.subjects
        ]
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(start_row, column, header)
            cell.font = Font(bold=True)
            cell.fill = self.AUTO_FILL
        for offset, values in enumerate(rows, start=1):
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(start_row + offset, column, value)
                cell.fill = self.AUTO_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        end_row = start_row + len(rows)
        table = Table(displayName="T_REFERENCE_MAP", ref=f"A{start_row}:D{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
        )
        worksheet.add_table(table)

    @staticmethod
    def _write_reference_guidance(workbook: Any) -> None:
        intro = workbook["00_最初に読む"]
        row = intro.max_row + 2
        intro.cell(row, 1, "・教師・教科は『入力用ラベル』で選択します。入力用ラベルは出力表示名とは別で、出力名が空欄の枠も一意に選択できます。")
        intro.cell(row + 1, 1, "・教師・教科のラベル文字列から意味を推測しないでください。内部IDとの対応は _system の参照マップが正本です。")
        for target_row in (row, row + 1):
            intro.cell(target_row, 1).alignment = Alignment(wrap_text=True, vertical="top")

        guide = workbook["90_AI編集ガイド"]
        row = guide.max_row + 2
        guide.cell(row, 1, "参照ラベル")
        guide.cell(row, 2, "教師・教科の入力用ラベル、内部ID、出力表示名は別概念。ラベルを解析せず _system/T_REFERENCE_MAP の明示対応を使う。")
        guide.cell(row + 1, 1, "出力名が空欄の枠")
        guide.cell(row + 1, 2, "入力Excelでは一意な入力用ラベルを持つが、出力表示名は空欄のまま。『(空欄)』という参照トークンでは識別しない。")
        for target_row in (row, row + 1):
            guide.cell(target_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            guide.cell(target_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _table(workbook: Any, sheet_name: str, table_name: str) -> tuple[Any, dict[str, int], int, int]:
        worksheet = workbook[sheet_name]
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = {
            str(worksheet.cell(min_row, column).value): column
            for column in range(min_col, max_col + 1)
        }
        return worksheet, headers, min_row, max_row
