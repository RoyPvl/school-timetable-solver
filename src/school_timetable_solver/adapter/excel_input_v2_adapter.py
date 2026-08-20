from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from school_timetable_solver.model.input_models import (
    CalendarDayModel,
    ClassPairOverlapRuleModel,
    HomeroomBoundaryRuleModel,
    InputDataModel,
    InputWorkbookSettingsModel,
    LessonCountPreferenceRuleSegmentModel,
    LessonCountRuleSegmentModel,
    LessonRequirementModel,
    PlacementRuleModel,
    TeacherDayOffRuleModel,
    TeacherLeaveModel,
)
from school_timetable_solver.model.master_models import (
    CampusModel,
    ClassModel,
    PeriodModel,
    RoomModel,
    SubjectModel,
    TeacherModel,
)
from school_timetable_solver.model.result_models import InputReadResultModel, ValidationIssueModel


YES = "○"
NO = "×"
UNSPECIFIED = "指定なし"
BLANK_TOKEN = "(空欄)"


class ExcelInputV2ReaderAdapter:
    """Read the self-describing, human-facing Excel input contract v2.0."""

    REQUIRED_TABLES = (
        "T_SETTINGS",
        "T_PERIODS",
        "T_CALENDAR",
        "T_CAMPUSES",
        "T_ROOMS",
        "T_TEACHERS",
        "T_CLASSES",
        "T_SUBJECTS",
        "T_LESSON_REQUIREMENTS",
        "T_TEACHER_LEAVES",
        "T_TEACHER_DAY_OFF_RULES",
        "T_PLACEMENT_RULES",
        "T_PLACEMENT_RULE_CONDITIONS",
        "T_LESSON_COUNT_HARD",
        "T_LESSON_COUNT_HARD_TARGETS",
        "T_LESSON_COUNT_SOFT",
        "T_LESSON_COUNT_SOFT_TARGETS",
        "T_HOMEROOM_RULES",
        "T_HOMEROOM_RULE_CONDITIONS",
        "T_CLASS_PAIRS",
        "T_DATE_SETS",
    )

    _division_to_internal = {
        "小学": "elementary",
        "中学": "junior_high",
        "高校": "high_school",
        "その他": "other",
    }
    _lesson_type_to_internal = {"通常": "regular", "特別": "special", "その他": "other"}
    _entity_to_internal = {"クラス": "class", "教師": "teacher"}
    _constraint_type_to_internal = {"追加制約": "hard", "上書き": "override"}
    _condition_field_to_internal = {
        "クラス": "class_id",
        "学部": "division",
        "学年": "grade",
        "受験区分": "exam_category",
        "校舎": "campus_id",
        "教師": "teacher_id",
        "担任通常授業あり": "has_regular_homeroom_lesson",
    }
    _operator_to_internal = {
        "=": "eq",
        "≠": "ne",
        "いずれか": "in",
        "以上": "ge",
        "以下": "le",
        "範囲": "between",
    }

    def read(self, path: Path) -> InputReadResultModel:
        issues: list[ValidationIssueModel] = []
        if not path.is_file():
            return self._result_error("INPUT_FILE_NOT_FOUND", str(path), "ファイルが存在しません")
        if path.suffix.lower() != ".xlsx":
            return self._result_error("INPUT_FILE_TYPE", str(path), ".xlsxファイルを指定してください")
        try:
            workbook = load_workbook(path, data_only=True)
        except (InvalidFileException, OSError, ValueError, KeyError) as exc:
            return self._result_error("INPUT_WORKBOOK_ERROR", str(path), f"Workbookを開けません: {exc}")

        try:
            tables = self._read_tables(workbook, issues)
            if self._has_errors(issues):
                return InputReadResultModel(None, tuple(issues))

            settings_rows = tables["T_SETTINGS"]
            settings_map = {self._text_or_blank(row.get("項目")): row.get("値") for row in settings_rows}
            timetable_name = self._required_text(settings_map.get("時間割名"), "T_SETTINGS", "時間割名", issues)
            description = self._optional_text(settings_map.get("説明"))

            periods = self._read_periods(tables["T_PERIODS"], issues)
            period_ids = tuple(item.period_id for item in sorted(periods, key=lambda item: item.output_order))
            period_columns = self._period_columns(periods)
            calendar_days = self._read_calendar(tables["T_CALENDAR"], period_columns, issues)
            campuses = self._read_campuses(tables["T_CAMPUSES"], issues)
            campus_by_name = self._unique_name_map(campuses, lambda item: item.campus_name, lambda item: item.campus_id, "校舎", issues)
            rooms = self._read_rooms(tables["T_ROOMS"], campus_by_name, issues)
            teachers = self._read_teachers(tables["T_TEACHERS"], campus_by_name, issues)
            teacher_by_name = self._unique_name_map(teachers, lambda item: item.teacher_name or BLANK_TOKEN, lambda item: item.teacher_id, "教師", issues)
            classes = self._read_classes(tables["T_CLASSES"], campus_by_name, teacher_by_name, issues)
            class_by_key = {(self._campus_name(campuses, item.campus_id), item.class_name): item.class_id for item in classes}
            subjects = self._read_subjects(tables["T_SUBJECTS"], issues)
            subject_by_name = self._unique_name_map(subjects, lambda item: item.subject_name or BLANK_TOKEN, lambda item: item.subject_id, "教科", issues)

            requirements = self._read_requirements(
                tables["T_LESSON_REQUIREMENTS"], class_by_key, subject_by_name, teacher_by_name, issues
            )
            teacher_leaves = self._read_teacher_leaves(
                tables["T_TEACHER_LEAVES"], teacher_by_name, period_columns, issues
            )
            date_sets = self._read_date_sets(tables["T_DATE_SETS"], issues)
            day_off_rules = self._read_day_off_rules(
                tables["T_TEACHER_DAY_OFF_RULES"], teacher_by_name, date_sets, issues
            )
            placement_conditions = self._group_conditions(
                tables["T_PLACEMENT_RULE_CONDITIONS"], "ルール", issues
            )
            placement_rules = self._read_placement_rules(
                tables["T_PLACEMENT_RULES"], placement_conditions, campus_by_name, class_by_key, teacher_by_name, period_columns, issues
            )
            hard_segments = self._read_lesson_count_rules(
                tables["T_LESSON_COUNT_HARD"],
                tables["T_LESSON_COUNT_HARD_TARGETS"],
                class_by_key,
                subject_by_name,
                period_columns,
                issues,
                soft=False,
            )
            soft_segments = self._read_lesson_count_rules(
                tables["T_LESSON_COUNT_SOFT"],
                tables["T_LESSON_COUNT_SOFT_TARGETS"],
                class_by_key,
                subject_by_name,
                period_columns,
                issues,
                soft=True,
            )
            homeroom_conditions = self._group_conditions(
                tables["T_HOMEROOM_RULE_CONDITIONS"], "ルール", issues
            )
            homeroom_rules = self._read_homeroom_rules(
                tables["T_HOMEROOM_RULES"], homeroom_conditions, class_by_key, teacher_by_name, campus_by_name, issues
            )
            class_pairs = self._read_class_pairs(tables["T_CLASS_PAIRS"], class_by_key, issues)

            if timetable_name is None or self._has_errors(issues):
                return InputReadResultModel(None, tuple(issues))
            return InputReadResultModel(
                InputDataModel(
                    settings=InputWorkbookSettingsModel("2.0", timetable_name, description),
                    calendar_days=tuple(calendar_days),
                    periods=tuple(periods),
                    campuses=tuple(campuses),
                    rooms=tuple(rooms),
                    teachers=tuple(teachers),
                    classes=tuple(classes),
                    subjects=tuple(subjects),
                    lesson_requirements=tuple(requirements),
                    teacher_leaves=tuple(teacher_leaves),
                    placement_rules=tuple(placement_rules),
                    lesson_count_rule_segments=tuple(hard_segments),
                    lesson_count_preference_rule_segments=tuple(soft_segments),
                    teacher_day_off_rules=tuple(day_off_rules),
                    homeroom_boundary_rules=tuple(homeroom_rules),
                    class_pair_overlap_rules=tuple(class_pairs),
                ),
                tuple(issues),
            )
        finally:
            workbook.close()

    def _read_tables(self, workbook: Any, issues: list[ValidationIssueModel]) -> dict[str, list[dict[str, Any]]]:
        found: dict[str, list[dict[str, Any]]] = {}
        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                name = table.name
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [worksheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
                rows: list[dict[str, Any]] = []
                for row_index in range(min_row + 1, max_row + 1):
                    values = [worksheet.cell(row_index, col).value for col in range(min_col, max_col + 1)]
                    if all(value is None for value in values):
                        continue
                    rows.append({str(header).strip(): value for header, value in zip(headers, values, strict=True) if header is not None})
                found[name] = rows
        for name in self.REQUIRED_TABLES:
            if name not in found:
                issues.append(self._issue("V2_TABLE_MISSING", name, f"必須の入力表がありません: {name}"))
        return found

    def _read_periods(self, rows: list[dict[str, Any]], issues: list[ValidationIssueModel]) -> list[PeriodModel]:
        result: list[PeriodModel] = []
        for index, row in enumerate(rows, start=1):
            period_id = self._optional_text(row.get("内部ID")) or f"P{index}"
            name = self._required_text(row.get("時限"), "T_PERIODS", f"行{index}", issues)
            order = self._required_int(row.get("表示順"), "T_PERIODS", f"行{index}", issues)
            start = self._required_time(row.get("開始"), "T_PERIODS", f"行{index}", issues)
            end = self._required_time(row.get("終了"), "T_PERIODS", f"行{index}", issues)
            if None not in (name, order, start, end):
                result.append(PeriodModel(period_id, name, order, start, end))
        return result

    def _period_columns(self, periods: list[PeriodModel]) -> tuple[tuple[str, str], ...]:
        return tuple((f"{index}限", period.period_id) for index, period in enumerate(sorted(periods, key=lambda item: item.output_order), start=1))

    def _read_calendar(self, rows: list[dict[str, Any]], period_columns: tuple[tuple[str, str], ...], issues: list[ValidationIssueModel]) -> list[CalendarDayModel]:
        result: list[CalendarDayModel] = []
        for index, row in enumerate(rows, start=1):
            target_date = self._required_date(row.get("日付"), "T_CALENDAR", f"行{index}", issues)
            output = self._required_yes_no(row.get("出力"), "T_CALENDAR", f"行{index}", issues)
            enabled: list[str] = []
            for column, period_id in period_columns:
                value = self._required_yes_no(row.get(column), "T_CALENDAR", f"行{index}/{column}", issues)
                if value:
                    enabled.append(period_id)
            if target_date is not None and output is not None:
                result.append(CalendarDayModel(target_date, output, tuple(enabled), self._optional_text(row.get("備考"))))
        return result

    def _read_campuses(self, rows: list[dict[str, Any]], issues: list[ValidationIssueModel]) -> list[CampusModel]:
        result: list[CampusModel] = []
        for index, row in enumerate(rows, start=1):
            name = self._required_text(row.get("校舎"), "T_CAMPUSES", f"行{index}", issues)
            order = self._required_int(row.get("表示順"), "T_CAMPUSES", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_CAMPUSES", f"行{index}", issues)
            if name is not None and order is not None and enabled is not None:
                result.append(CampusModel(self._optional_text(row.get("内部ID")) or f"CAMPUS_V2_{index:03d}", name, order, enabled))
        return result

    def _read_rooms(self, rows: list[dict[str, Any]], campus_by_name: dict[str, str], issues: list[ValidationIssueModel]) -> list[RoomModel]:
        result: list[RoomModel] = []
        for index, row in enumerate(rows, start=1):
            campus_id = self._resolve_name(row.get("校舎"), campus_by_name, "校舎", "T_ROOMS", index, issues)
            name = self._required_text(row.get("教室"), "T_ROOMS", f"行{index}", issues)
            order = self._required_int(row.get("表示順"), "T_ROOMS", f"行{index}", issues)
            priority = self._required_int(row.get("優先度"), "T_ROOMS", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_ROOMS", f"行{index}", issues)
            if None not in (campus_id, name, order, priority, enabled):
                result.append(RoomModel(self._optional_text(row.get("内部ID")) or f"ROOM_V2_{index:03d}", name, campus_id, order, priority, enabled))
        return result

    def _read_teachers(self, rows: list[dict[str, Any]], campus_by_name: dict[str, str], issues: list[ValidationIssueModel]) -> list[TeacherModel]:
        result: list[TeacherModel] = []
        for index, row in enumerate(rows, start=1):
            display = self._required_text(row.get("教師"), "T_TEACHERS", f"行{index}", issues)
            campus_id = self._resolve_name(row.get("主校舎"), campus_by_name, "校舎", "T_TEACHERS", index, issues)
            enabled = self._required_yes_no(row.get("使用"), "T_TEACHERS", f"行{index}", issues)
            if display is not None and campus_id is not None and enabled is not None:
                internal_name = self._optional_text(row.get("内部表示名"))
                name = internal_name if display == BLANK_TOKEN and internal_name is not None else ("" if display == BLANK_TOKEN else display)
                result.append(TeacherModel(self._optional_text(row.get("内部ID")) or f"TEACHER_V2_{index:03d}", name, campus_id, enabled))
        return result

    def _read_classes(self, rows: list[dict[str, Any]], campus_by_name: dict[str, str], teacher_by_name: dict[str, str], issues: list[ValidationIssueModel]) -> list[ClassModel]:
        result: list[ClassModel] = []
        for index, row in enumerate(rows, start=1):
            campus_id = self._resolve_name(row.get("校舎"), campus_by_name, "校舎", "T_CLASSES", index, issues)
            name = self._required_text(row.get("クラス"), "T_CLASSES", f"行{index}", issues)
            division_display = self._required_text(row.get("学部"), "T_CLASSES", f"行{index}", issues)
            division = self._division_to_internal.get(division_display or "", self._optional_text(row.get("内部学部")) or division_display)
            grade = self._required_int(row.get("学年"), "T_CLASSES", f"行{index}", issues)
            exam_category = self._required_text(row.get("区分"), "T_CLASSES", f"行{index}", issues)
            homeroom_display = self._optional_text(row.get("担任"))
            homeroom_id = None if homeroom_display in (None, UNSPECIFIED, BLANK_TOKEN) else self._resolve_name(homeroom_display, teacher_by_name, "教師", "T_CLASSES", index, issues)
            enabled = self._required_yes_no(row.get("使用"), "T_CLASSES", f"行{index}", issues)
            if None not in (campus_id, name, division, grade, exam_category, enabled):
                result.append(ClassModel(self._optional_text(row.get("内部ID")) or f"CLASS_V2_{index:03d}", name, campus_id, str(division), grade, exam_category, homeroom_id, enabled))
        return result

    def _read_subjects(self, rows: list[dict[str, Any]], issues: list[ValidationIssueModel]) -> list[SubjectModel]:
        result: list[SubjectModel] = []
        for index, row in enumerate(rows, start=1):
            display = self._required_text(row.get("教科"), "T_SUBJECTS", f"行{index}", issues)
            type_display = self._required_text(row.get("種別"), "T_SUBJECTS", f"行{index}", issues)
            lesson_type = self._lesson_type_to_internal.get(type_display or "", self._optional_text(row.get("内部種別")) or type_display)
            enabled = self._required_yes_no(row.get("使用"), "T_SUBJECTS", f"行{index}", issues)
            if display is not None and lesson_type is not None and enabled is not None:
                internal_name = self._optional_text(row.get("内部表示名"))
                name = internal_name if display == BLANK_TOKEN and internal_name is not None else ("" if display == BLANK_TOKEN else display)
                result.append(SubjectModel(self._optional_text(row.get("内部ID")) or f"SUBJECT_V2_{index:03d}", name, str(lesson_type), enabled))
        return result

    def _read_requirements(self, rows: list[dict[str, Any]], class_by_key: dict[tuple[str, str], str], subject_by_name: dict[str, str], teacher_by_name: dict[str, str], issues: list[ValidationIssueModel]) -> list[LessonRequirementModel]:
        result: list[LessonRequirementModel] = []
        for index, row in enumerate(rows, start=1):
            class_id = self._resolve_class(row, class_by_key, "T_LESSON_REQUIREMENTS", index, issues)
            subject_id = self._resolve_name(row.get("教科"), subject_by_name, "教科", "T_LESSON_REQUIREMENTS", index, issues)
            teacher_id = self._resolve_name(row.get("教師"), teacher_by_name, "教師", "T_LESSON_REQUIREMENTS", index, issues)
            required = self._required_int(row.get("必要コマ"), "T_LESSON_REQUIREMENTS", f"行{index}", issues)
            maximum = self._optional_int(row.get("1日最大"), "T_LESSON_REQUIREMENTS", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_LESSON_REQUIREMENTS", f"行{index}", issues)
            if None not in (class_id, subject_id, teacher_id, required, enabled):
                result.append(LessonRequirementModel(self._optional_text(row.get("内部ID")) or f"REQ_V2_{index:04d}", class_id, subject_id, teacher_id, required, maximum, enabled))
        return result

    def _read_teacher_leaves(self, rows: list[dict[str, Any]], teacher_by_name: dict[str, str], period_columns: tuple[tuple[str, str], ...], issues: list[ValidationIssueModel]) -> list[TeacherLeaveModel]:
        result: list[TeacherLeaveModel] = []
        for index, row in enumerate(rows, start=1):
            teacher_id = self._resolve_name(row.get("教師"), teacher_by_name, "教師", "T_TEACHER_LEAVES", index, issues)
            target_date = self._required_date(row.get("日付"), "T_TEACHER_LEAVES", f"行{index}", issues)
            unavailable: list[str] = []
            for column, period_id in period_columns:
                value = self._required_yes_no(row.get(f"{column}休み"), "T_TEACHER_LEAVES", f"行{index}/{column}", issues)
                if value:
                    unavailable.append(period_id)
            if teacher_id is not None and target_date is not None:
                result.append(TeacherLeaveModel(teacher_id, target_date, tuple(unavailable)))
        return result

    def _read_date_sets(self, rows: list[dict[str, Any]], issues: list[ValidationIssueModel]) -> dict[str, tuple[date, ...]]:
        grouped: dict[str, list[date]] = defaultdict(list)
        for index, row in enumerate(rows, start=1):
            name = self._required_text(row.get("日付集合"), "T_DATE_SETS", f"行{index}", issues)
            value = self._required_date(row.get("日付"), "T_DATE_SETS", f"行{index}", issues)
            if name is not None and value is not None:
                grouped[name].append(value)
        return {name: tuple(dict.fromkeys(values)) for name, values in grouped.items()}

    def _read_day_off_rules(self, rows: list[dict[str, Any]], teacher_by_name: dict[str, str], date_sets: dict[str, tuple[date, ...]], issues: list[ValidationIssueModel]) -> list[TeacherDayOffRuleModel]:
        result: list[TeacherDayOffRuleModel] = []
        for index, row in enumerate(rows, start=1):
            teacher_id = self._resolve_name(row.get("教師"), teacher_by_name, "教師", "T_TEACHER_DAY_OFF_RULES", index, issues)
            set_name = self._required_text(row.get("候補日集合"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues)
            eligible_dates = date_sets.get(set_name or "")
            if set_name is not None and eligible_dates is None:
                issues.append(self._issue("V2_DATE_SET_NOT_FOUND", set_name, f"日付集合が見つかりません: {set_name}"))
            enabled = self._required_yes_no(row.get("使用"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues)
            if teacher_id is not None and eligible_dates is not None and enabled is not None:
                result.append(TeacherDayOffRuleModel(
                    rule_id=self._optional_text(row.get("内部ID")) or f"DAYOFF_V2_{index:03d}",
                    teacher_id=teacher_id,
                    enabled=enabled,
                    eligible_dates=eligible_dates,
                    required_days_off=self._optional_int(row.get("必須休み日数"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues),
                    minimum_days_off=self._optional_int(row.get("最低休み日数"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues),
                    maximum_days_off=self._optional_int(row.get("最大休み日数"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues),
                    preferred_days_off=self._optional_int(row.get("希望休み日数"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues),
                    quota_group_id=self._none_if_unspecified(row.get("グループ")),
                    group_required_days_off=self._optional_int(row.get("グループ必須休み日数"), "T_TEACHER_DAY_OFF_RULES", f"行{index}", issues),
                ))
        return result

    def _group_conditions(self, rows: list[dict[str, Any]], rule_column: str, issues: list[ValidationIssueModel]) -> dict[str, list[tuple[str, str, str]]]:
        grouped: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
        for index, row in enumerate(rows, start=1):
            rule = self._required_text(row.get(rule_column), "CONDITIONS", f"行{index}", issues)
            field_display = self._required_text(row.get("条件項目"), "CONDITIONS", f"行{index}", issues)
            operator_display = self._required_text(row.get("比較"), "CONDITIONS", f"行{index}", issues)
            value = self._required_text(row.get("値"), "CONDITIONS", f"行{index}", issues)
            field = self._condition_field_to_internal.get(field_display or "", self._optional_text(row.get("内部条件項目")) or field_display)
            operator = self._operator_to_internal.get(operator_display or "", self._optional_text(row.get("内部比較")) or operator_display)
            if None not in (rule, field, operator, value):
                grouped[rule].append((str(field), str(operator), value))
        return grouped

    def _read_placement_rules(self, rows: list[dict[str, Any]], conditions: dict[str, list[tuple[str, str, str]]], campus_by_name: dict[str, str], class_by_key: dict[tuple[str, str], str], teacher_by_name: dict[str, str], period_columns: tuple[tuple[str, str], ...], issues: list[ValidationIssueModel]) -> list[PlacementRuleModel]:
        result: list[PlacementRuleModel] = []
        for index, row in enumerate(rows, start=1):
            name = self._required_text(row.get("ルール"), "T_PLACEMENT_RULES", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_PLACEMENT_RULES", f"行{index}", issues)
            entity_display = self._required_text(row.get("対象"), "T_PLACEMENT_RULES", f"行{index}", issues)
            entity = self._entity_to_internal.get(entity_display or "", entity_display)
            type_display = self._required_text(row.get("種別"), "T_PLACEMENT_RULES", f"行{index}", issues)
            constraint_type = self._constraint_type_to_internal.get(type_display or "", type_display)
            campus_display = self._none_if_unspecified(row.get("校舎"))
            campus_id = None if campus_display is None else self._resolve_name(campus_display, campus_by_name, "校舎", "T_PLACEMENT_RULES", index, issues)
            period_specified = self._required_yes_no(row.get("時限を指定"), "T_PLACEMENT_RULES", f"行{index}", issues)
            allowed: list[str] = []
            required: list[str] = []
            if period_specified:
                for column, period_id in period_columns:
                    if self._required_yes_no(row.get(f"{column}許可"), "T_PLACEMENT_RULES", f"行{index}/{column}許可", issues):
                        allowed.append(period_id)
            for column, period_id in period_columns:
                required_value = self._optional_yes_no(row.get(f"{column}必須"), "T_PLACEMENT_RULES", f"行{index}/{column}必須", issues)
                if required_value:
                    required.append(period_id)
            rule_conditions = list(conditions.get(name or "", ()))
            direct_class = self._none_if_unspecified(row.get("クラス"))
            if direct_class is not None:
                class_id = self._resolve_class(row, class_by_key, "T_PLACEMENT_RULES", index, issues)
                if class_id is not None:
                    rule_conditions.append(("class_id", "eq", class_id))
            direct_teacher = self._none_if_unspecified(row.get("教師"))
            if direct_teacher is not None:
                teacher_id = self._resolve_name(direct_teacher, teacher_by_name, "教師", "T_PLACEMENT_RULES", index, issues)
                if teacher_id is not None:
                    rule_conditions.append(("teacher_id", "eq", teacher_id))
            if name is not None and enabled is not None and entity is not None and constraint_type is not None:
                result.append(PlacementRuleModel(
                    rule_id=self._optional_text(row.get("内部ID")) or f"PLACEMENT_V2_{index:03d}",
                    rule_name=name,
                    enabled=enabled,
                    constraint_type=str(constraint_type),
                    target_entity=str(entity),
                    condition_fields=tuple(item[0] for item in rule_conditions),
                    condition_operators=tuple(item[1] for item in rule_conditions),
                    condition_values=tuple(item[2] for item in rule_conditions),
                    campus_id=campus_id,
                    start_date=self._optional_date(row.get("開始日"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    end_date=self._optional_date(row.get("終了日"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    weekdays=tuple(self._split_tokens(row.get("曜日"))),
                    allowed_period_ids=tuple(allowed),
                    daily_hard_limit=self._optional_int(row.get("1日上限"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    forbid_first_last_same_day=self._optional_yes_no(row.get("1限6限同日禁止"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    attendance_streak_limit=self._optional_int(row.get("連続登校上限"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    priority=self._required_int(row.get("優先度"), "T_PLACEMENT_RULES", f"行{index}", issues) or 0,
                    preferred_attendance_streak_limit=self._optional_int(row.get("希望連続登校上限"), "T_PLACEMENT_RULES", f"行{index}", issues),
                    required_lesson_period_ids=tuple(required),
                ))
        return result

    def _read_lesson_count_rules(self, rule_rows: list[dict[str, Any]], target_rows: list[dict[str, Any]], class_by_key: dict[tuple[str, str], str], subject_by_name: dict[str, str], period_columns: tuple[tuple[str, str], ...], issues: list[ValidationIssueModel], *, soft: bool) -> list[LessonCountRuleSegmentModel] | list[LessonCountPreferenceRuleSegmentModel]:
        table = "T_LESSON_COUNT_SOFT" if soft else "T_LESSON_COUNT_HARD"
        target_table = f"{table}_TARGETS"
        targets: dict[str, list[tuple[str, str]]] = defaultdict(list)
        for index, row in enumerate(target_rows, start=1):
            rule = self._required_text(row.get("ルール"), target_table, f"行{index}", issues)
            class_id = self._resolve_class(row, class_by_key, target_table, index, issues)
            subject_id = self._resolve_name(row.get("教科"), subject_by_name, "教科", target_table, index, issues)
            if None not in (rule, class_id, subject_id):
                targets[rule].append((class_id, subject_id))
        segments_by_rule: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rule_rows:
            rule = self._required_text(row.get("ルール"), table, "ルール", issues)
            if rule is not None:
                segments_by_rule[rule].append(row)
        result: list[Any] = []
        for rule_name, segments in segments_by_rule.items():
            if not targets.get(rule_name):
                issues.append(self._issue("V2_RULE_TARGET_EMPTY", rule_name, f"対象がありません: {rule_name}"))
                continue
            for target_index, (class_id, subject_id) in enumerate(targets[rule_name], start=1):
                generated_rule_id = f"{'SOFT' if soft else 'HARD'}_V2_{self._safe_id(rule_name)}_{target_index:03d}"
                for segment_index, row in enumerate(segments, start=1):
                    enabled = self._required_yes_no(row.get("使用"), table, rule_name, issues)
                    count_column = "希望配置数" if soft else "配置数"
                    count = self._required_int(row.get(count_column), table, rule_name, issues)
                    start_date = self._required_date(row.get("開始日"), table, rule_name, issues)
                    end_date = self._required_date(row.get("終了日"), table, rule_name, issues)
                    target_periods = tuple(period_id for column, period_id in period_columns if self._required_yes_no(row.get(f"{column}対象"), table, f"{rule_name}/{column}", issues))
                    if None in (enabled, count, start_date, end_date):
                        continue
                    if soft:
                        result.append(LessonCountPreferenceRuleSegmentModel(generated_rule_id, f"SEG{segment_index:02d}", rule_name, enabled, class_id, subject_id, count, start_date, end_date, target_periods))
                    else:
                        result.append(LessonCountRuleSegmentModel(generated_rule_id, f"SEG{segment_index:02d}", rule_name, enabled, class_id, subject_id, count, start_date, end_date, target_periods))
        return result

    def _read_homeroom_rules(self, rows: list[dict[str, Any]], conditions: dict[str, list[tuple[str, str, str]]], class_by_key: dict[tuple[str, str], str], teacher_by_name: dict[str, str], campus_by_name: dict[str, str], issues: list[ValidationIssueModel]) -> list[HomeroomBoundaryRuleModel]:
        result: list[HomeroomBoundaryRuleModel] = []
        for index, row in enumerate(rows, start=1):
            name = self._required_text(row.get("ルール"), "T_HOMEROOM_RULES", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_HOMEROOM_RULES", f"行{index}", issues)
            start = self._required_date(row.get("開始日"), "T_HOMEROOM_RULES", f"行{index}", issues)
            end = self._required_date(row.get("終了日"), "T_HOMEROOM_RULES", f"行{index}", issues)
            rule_conditions = conditions.get(name or "", [])
            if None not in (name, enabled, start, end):
                result.append(HomeroomBoundaryRuleModel(self._optional_text(row.get("内部ID")) or f"HOMEROOM_V2_{index:03d}", name, enabled, tuple(item[0] for item in rule_conditions), tuple(item[1] for item in rule_conditions), tuple(item[2] for item in rule_conditions), start, end))
        return result

    def _read_class_pairs(self, rows: list[dict[str, Any]], class_by_key: dict[tuple[str, str], str], issues: list[ValidationIssueModel]) -> list[ClassPairOverlapRuleModel]:
        result: list[ClassPairOverlapRuleModel] = []
        for index, row in enumerate(rows, start=1):
            name = self._required_text(row.get("ルール"), "T_CLASS_PAIRS", f"行{index}", issues)
            enabled = self._required_yes_no(row.get("使用"), "T_CLASS_PAIRS", f"行{index}", issues)
            first = self._resolve_class({"校舎": row.get("校舎A"), "クラス": row.get("クラスA")}, class_by_key, "T_CLASS_PAIRS", index, issues)
            second = self._resolve_class({"校舎": row.get("校舎B"), "クラス": row.get("クラスB")}, class_by_key, "T_CLASS_PAIRS", index, issues)
            if None not in (name, enabled, first, second):
                result.append(ClassPairOverlapRuleModel(self._optional_text(row.get("内部ID")) or f"PAIR_V2_{index:03d}", name, enabled, first, second))
        return result

    def _resolve_class(self, row: dict[str, Any], class_by_key: dict[tuple[str, str], str], table: str, index: int, issues: list[ValidationIssueModel]) -> str | None:
        campus = self._required_text(row.get("校舎"), table, f"行{index}/校舎", issues)
        class_name = self._required_text(row.get("クラス"), table, f"行{index}/クラス", issues)
        if campus is None or class_name is None:
            return None
        value = class_by_key.get((campus, class_name))
        if value is None:
            issues.append(self._issue("V2_CLASS_NOT_FOUND", f"{campus}/{class_name}", f"クラスが見つかりません: {campus} / {class_name}"))
        return value

    def _resolve_name(self, value: Any, mapping: dict[str, str], label: str, table: str, index: int, issues: list[ValidationIssueModel]) -> str | None:
        text = self._required_text(value, table, f"行{index}/{label}", issues)
        if text is None:
            return None
        result = mapping.get(text)
        if result is None:
            issues.append(self._issue("V2_REFERENCE_NOT_FOUND", text, f"{label}一覧に「{text}」が見つかりません"))
        return result

    def _unique_name_map(self, items: list[Any], name_getter: Any, id_getter: Any, label: str, issues: list[ValidationIssueModel]) -> dict[str, str]:
        result: dict[str, str] = {}
        for item in items:
            name = name_getter(item)
            if name in result:
                issues.append(self._issue("V2_DUPLICATE_DISPLAY_NAME", name, f"{label}名が重複しているため一意に参照できません: {name}"))
            result[name] = id_getter(item)
        return result

    @staticmethod
    def _campus_name(campuses: list[CampusModel], campus_id: str) -> str:
        return next(item.campus_name for item in campuses if item.campus_id == campus_id)

    def _required_yes_no(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> bool | None:
        if value == YES:
            return True
        if value == NO:
            return False
        issues.append(self._issue("V2_YES_NO_REQUIRED", target, f"{table}: 「○」または「×」を選択してください"))
        return None

    def _optional_yes_no(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> bool | None:
        if value in (None, "", UNSPECIFIED):
            return None
        return self._required_yes_no(value, table, target, issues)

    def _required_text(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> str | None:
        text = self._optional_text(value)
        if text is None:
            issues.append(self._issue("V2_REQUIRED_VALUE", target, f"{table}: 必須項目が空です"))
        return text

    @staticmethod
    def _optional_text(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _text_or_blank(value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _required_int(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> int | None:
        parsed = self._optional_int(value, table, target, issues)
        if parsed is None:
            issues.append(self._issue("V2_INTEGER_REQUIRED", target, f"{table}: 整数を入力してください"))
        return parsed

    def _optional_int(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> int | None:
        if value in (None, "", UNSPECIFIED):
            return None
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            issues.append(self._issue("V2_INTEGER", target, f"{table}: 整数として読めません: {value}"))
            return None
        return parsed

    def _required_date(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> date | None:
        parsed = self._optional_date(value, table, target, issues)
        if parsed is None:
            issues.append(self._issue("V2_DATE_REQUIRED", target, f"{table}: 日付を入力してください"))
        return parsed

    def _optional_date(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> date | None:
        if value in (None, "", UNSPECIFIED):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            issues.append(self._issue("V2_DATE", target, f"{table}: 日付として読めません: {value}"))
            return None

    def _required_time(self, value: Any, table: str, target: str, issues: list[ValidationIssueModel]) -> time | None:
        if isinstance(value, datetime):
            return value.time()
        if isinstance(value, time):
            return value
        if isinstance(value, str):
            try:
                return time.fromisoformat(value)
            except ValueError:
                pass
        issues.append(self._issue("V2_TIME", target, f"{table}: 時刻として読めません: {value}"))
        return None

    @staticmethod
    def _none_if_unspecified(value: Any) -> str | None:
        if value in (None, "", UNSPECIFIED):
            return None
        return str(value).strip()

    @staticmethod
    def _split_tokens(value: Any) -> list[str]:
        if value in (None, "", UNSPECIFIED):
            return []
        return [token.strip() for token in str(value).replace(",", "/").split("/") if token.strip()]

    @staticmethod
    def _safe_id(value: str) -> str:
        return "".join(character if character.isalnum() else "_" for character in value)[:40]

    @staticmethod
    def _has_errors(issues: list[ValidationIssueModel]) -> bool:
        return any(issue.severity == "ERROR" for issue in issues)

    @staticmethod
    def _issue(rule_id: str, target: str, message: str) -> ValidationIssueModel:
        return ValidationIssueModel(rule_id, "ERROR", target, message, None)

    def _result_error(self, rule_id: str, target: str, message: str) -> InputReadResultModel:
        return InputReadResultModel(None, (self._issue(rule_id, target, message),))
