from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from school_timetable_solver.adapter.excel_input_adapter import ExcelInputReaderAdapter
from school_timetable_solver.adapter.excel_input_v2_entity_reference_adapter import (
    EntityReferenceExcelInputV2ReaderAdapter,
)
from school_timetable_solver.model.result_models import InputReadResultModel, ValidationIssueModel


class CompatibleExcelInputReaderAdapter:
    """Route v1.1 and v2.0 workbooks without changing downstream models."""

    def __init__(self) -> None:
        self._v1_reader = ExcelInputReaderAdapter()
        self._v2_reader = EntityReferenceExcelInputV2ReaderAdapter()

    def read(self, path: Path) -> InputReadResultModel:
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            return self._v1_reader.read(path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError, KeyError):
            return self._v1_reader.read(path)
        try:
            if "_system" in workbook.sheetnames:
                system_sheet = workbook["_system"]
                for row in system_sheet.iter_rows(values_only=True):
                    if len(row) >= 2 and row[0] == "schema_version" and str(row[1]).strip() == "2.0":
                        return self._v2_reader.read(path)
        finally:
            workbook.close()
        return self._v1_reader.read(path)

    @staticmethod
    def unsupported_v2_issue(path: Path, message: str) -> InputReadResultModel:
        return InputReadResultModel(
            None,
            (
                ValidationIssueModel(
                    "INPUT_V2_ERROR",
                    "ERROR",
                    str(path),
                    message,
                    None,
                ),
            ),
        )
