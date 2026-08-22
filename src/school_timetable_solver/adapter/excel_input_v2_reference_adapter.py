from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from school_timetable_solver.adapter.excel_input_v2_adapter import (
    BLANK_TOKEN,
    ExcelInputV2ReaderAdapter,
)
from school_timetable_solver.model.master_models import SubjectModel, TeacherModel
from school_timetable_solver.model.result_models import InputReadResultModel, ValidationIssueModel


class ReferenceLabelExcelInputV2ReaderAdapter(ExcelInputV2ReaderAdapter):
    """Read v2 workbooks whose human input labels are separated from output names.

    New v2 migrations contain ``T_REFERENCE_MAP`` in ``_system``. The map is the
    authoritative bridge from a human-facing input label to an internal entity ID
    and its output display name. No special label text is parsed or interpreted.

    Workbooks created by the first v2 implementation did not contain this table;
    those files continue to use the original reader behavior for compatibility.
    """

    def __init__(self) -> None:
        self._teacher_label_to_id: dict[str, str] | None = None
        self._subject_label_to_id: dict[str, str] | None = None
        self._teacher_output_name_by_id: dict[str, str] = {}
        self._subject_output_name_by_id: dict[str, str] = {}

    def read(self, path: Path) -> InputReadResultModel:
        self._reset_reference_maps()
        issues: list[ValidationIssueModel] = []
        if path.is_file() and path.suffix.lower() == ".xlsx":
            try:
                workbook = load_workbook(path, data_only=True)
            except (OSError, ValueError, KeyError):
                workbook = None
            if workbook is not None:
                try:
                    self._load_reference_map(workbook, issues)
                finally:
                    workbook.close()
        if issues:
            return InputReadResultModel(None, tuple(issues))
        return super().read(path)

    def _reset_reference_maps(self) -> None:
        self._teacher_label_to_id = None
        self._subject_label_to_id = None
        self._teacher_output_name_by_id = {}
        self._subject_output_name_by_id = {}

    def _load_reference_map(self, workbook: Any, issues: list[ValidationIssueModel]) -> None:
        table = None
        for worksheet in workbook.worksheets:
            if "T_REFERENCE_MAP" in worksheet.tables:
                table = worksheet.tables["T_REFERENCE_MAP"]
                table_sheet = worksheet
                break
        if table is None:
            return

        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = {
            str(table_sheet.cell(min_row, column).value).strip(): column
            for column in range(min_col, max_col + 1)
        }
        required = {"参照種別", "内部ID", "入力用ラベル", "出力表示名"}
        missing = required.difference(headers)
        if missing:
            issues.append(
                self._issue(
                    "V2_REFERENCE_MAP_HEADER",
                    "T_REFERENCE_MAP",
                    f"参照マップの必須列がありません: {', '.join(sorted(missing))}",
                )
            )
            return

        teacher_labels: dict[str, str] = {}
        subject_labels: dict[str, str] = {}
        teacher_names: dict[str, str] = {}
        subject_names: dict[str, str] = {}
        for row_index in range(min_row + 1, max_row + 1):
            entity = self._optional_text(table_sheet.cell(row_index, headers["参照種別"]).value)
            internal_id = self._optional_text(table_sheet.cell(row_index, headers["内部ID"]).value)
            label = self._optional_text(table_sheet.cell(row_index, headers["入力用ラベル"]).value)
            raw_output = table_sheet.cell(row_index, headers["出力表示名"]).value
            output_name = "" if raw_output is None else str(raw_output)
            if entity is None and internal_id is None and label is None:
                continue
            if entity not in {"teacher", "subject"} or internal_id is None or label is None:
                issues.append(
                    self._issue(
                        "V2_REFERENCE_MAP_ROW",
                        f"T_REFERENCE_MAP!{row_index}",
                        "参照種別・内部ID・入力用ラベルを正しく入力してください",
                    )
                )
                continue
            target_labels = teacher_labels if entity == "teacher" else subject_labels
            target_names = teacher_names if entity == "teacher" else subject_names
            if label in target_labels:
                issues.append(
                    self._issue(
                        "V2_DUPLICATE_INPUT_LABEL",
                        label,
                        f"入力用ラベルが重複しています: {label}",
                    )
                )
                continue
            if internal_id in target_names:
                issues.append(
                    self._issue(
                        "V2_DUPLICATE_REFERENCE_ID",
                        internal_id,
                        f"参照マップの内部IDが重複しています: {internal_id}",
                    )
                )
                continue
            target_labels[label] = internal_id
            target_names[internal_id] = output_name

        self._teacher_label_to_id = teacher_labels
        self._subject_label_to_id = subject_labels
        self._teacher_output_name_by_id = teacher_names
        self._subject_output_name_by_id = subject_names

    def _read_teachers(
        self,
        rows: list[dict[str, Any]],
        campus_by_name: dict[str, str],
        issues: list[ValidationIssueModel],
    ) -> list[TeacherModel]:
        if self._teacher_label_to_id is None:
            return super()._read_teachers(rows, campus_by_name, issues)
        result: list[TeacherModel] = []
        for index, row in enumerate(rows, start=1):
            internal_id = self._required_text(row.get("内部ID"), "T_TEACHERS", f"行{index}/内部ID", issues)
            label = self._required_text(row.get("教師"), "T_TEACHERS", f"行{index}/教師", issues)
            campus_id = self._resolve_name(row.get("主校舎"), campus_by_name, "校舎", "T_TEACHERS", index, issues)
            enabled = self._required_yes_no(row.get("使用"), "T_TEACHERS", f"行{index}", issues)
            if internal_id is None or label is None or campus_id is None or enabled is None:
                continue
            mapped_id = self._teacher_label_to_id.get(label)
            if mapped_id != internal_id:
                issues.append(
                    self._issue(
                        "V2_REFERENCE_LABEL_MISMATCH",
                        label,
                        f"教師の入力用ラベルと内部IDの対応が参照マップと一致しません: {label}",
                    )
                )
                continue
            if internal_id not in self._teacher_output_name_by_id:
                issues.append(self._issue("V2_REFERENCE_NOT_FOUND", internal_id, f"教師参照が見つかりません: {internal_id}"))
                continue
            result.append(
                TeacherModel(
                    internal_id,
                    self._teacher_output_name_by_id[internal_id],
                    campus_id,
                    enabled,
                )
            )
        return result

    def _read_subjects(
        self,
        rows: list[dict[str, Any]],
        issues: list[ValidationIssueModel],
    ) -> list[SubjectModel]:
        if self._subject_label_to_id is None:
            return super()._read_subjects(rows, issues)
        result: list[SubjectModel] = []
        for index, row in enumerate(rows, start=1):
            internal_id = self._required_text(row.get("内部ID"), "T_SUBJECTS", f"行{index}/内部ID", issues)
            label = self._required_text(row.get("教科"), "T_SUBJECTS", f"行{index}/教科", issues)
            type_display = self._required_text(row.get("種別"), "T_SUBJECTS", f"行{index}", issues)
            lesson_type = self._lesson_type_to_internal.get(
                type_display or "",
                self._optional_text(row.get("内部種別")) or type_display,
            )
            enabled = self._required_yes_no(row.get("使用"), "T_SUBJECTS", f"行{index}", issues)
            if internal_id is None or label is None or lesson_type is None or enabled is None:
                continue
            mapped_id = self._subject_label_to_id.get(label)
            if mapped_id != internal_id:
                issues.append(
                    self._issue(
                        "V2_REFERENCE_LABEL_MISMATCH",
                        label,
                        f"教科の入力用ラベルと内部IDの対応が参照マップと一致しません: {label}",
                    )
                )
                continue
            if internal_id not in self._subject_output_name_by_id:
                issues.append(self._issue("V2_REFERENCE_NOT_FOUND", internal_id, f"教科参照が見つかりません: {internal_id}"))
                continue
            result.append(
                SubjectModel(
                    internal_id,
                    self._subject_output_name_by_id[internal_id],
                    str(lesson_type),
                    enabled,
                )
            )
        return result

    def _unique_name_map(
        self,
        items: list[Any],
        name_getter: Any,
        id_getter: Any,
        label: str,
        issues: list[ValidationIssueModel],
    ) -> dict[str, str]:
        if label == "教師" and self._teacher_label_to_id is not None:
            item_ids = {id_getter(item) for item in items}
            missing = set(self._teacher_label_to_id.values()).difference(item_ids)
            for internal_id in sorted(missing):
                issues.append(self._issue("V2_REFERENCE_NOT_FOUND", internal_id, f"教師マスタに内部IDがありません: {internal_id}"))
            return dict(self._teacher_label_to_id)
        if label == "教科" and self._subject_label_to_id is not None:
            item_ids = {id_getter(item) for item in items}
            missing = set(self._subject_label_to_id.values()).difference(item_ids)
            for internal_id in sorted(missing):
                issues.append(self._issue("V2_REFERENCE_NOT_FOUND", internal_id, f"教科マスタに内部IDがありません: {internal_id}"))
            return dict(self._subject_label_to_id)
        return super()._unique_name_map(items, name_getter, id_getter, label, issues)
