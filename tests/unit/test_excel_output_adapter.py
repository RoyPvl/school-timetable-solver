from __future__ import annotations

from dataclasses import replace
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from school_timetable_solver.adapter.excel_output_adapter import (
    ExcelTimetableWriterAdapter,
)
from school_timetable_solver.model.input_models import InputDataModel, TeacherLeaveModel
from school_timetable_solver.model.result_models import ScheduledLessonModel
from school_timetable_solver.service.result_services import BuildTimetableDocumentService


def _excel_date(value: object) -> date:
    assert isinstance(value, datetime)
    return value.date()


def test_writer_generates_contractual_matrix_layout_and_borders(
    minimal_input_data: InputDataModel,
    tmp_path: Path,
) -> None:
    lesson = ScheduledLessonModel(
        "Q1",
        minimal_input_data.calendar_days[0].target_date,
        "P1",
        "T1",
        "R1",
        "C1",
        "CL1",
        "S1",
    )
    document = BuildTimetableDocumentService().execute(minimal_input_data, (lesson,))
    output = tmp_path / "result.xlsx"
    ExcelTimetableWriterAdapter().write(document, output)

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook["全体"]
    assert workbook.sheetnames == ["全体"]
    assert worksheet["A1"].value is None
    assert _excel_date(worksheet["A2"].value) == minimal_input_data.calendar_days[0].target_date
    assert worksheet["A2"].number_format == 'm"月"d"日"'
    assert worksheet["A3"].value == "月曜日"
    assert worksheet["A4"].value == "①\n9:00～10:00"  # noqa: RUF001
    assert worksheet["B4"].value == "クラス"
    assert worksheet["C4"].value == "小学A"
    assert worksheet["C5"].value == "算数"
    assert worksheet["C6"].value == "教師一"
    assert {"A2:B2", "A3:B3", "C2:D2", "A4:A6"} <= {
        str(item) for item in worksheet.merged_cells.ranges
    }
    assert worksheet["A2"].border.top.style == "thin"
    assert worksheet["A2"].border.bottom.style == "hair"
    assert worksheet["A2"].border.right.style == "thin"
    assert worksheet["C4"].border.top.style == "thin"
    assert worksheet["C4"].border.bottom.style == "hair"
    assert worksheet["C4"].border.right.style == "hair"
    assert worksheet["D4"].border.right.style == "thin"
    assert worksheet["E4"].border.left.style == "thin"
    assert worksheet["A4"].border.bottom.style == "thin"
    assert workbook.__dict__.get("_external_links") == []


def test_writer_outputs_empty_subject_and_teacher_names_as_blank_cells(
    minimal_input_data: InputDataModel,
    tmp_path: Path,
) -> None:
    input_data = replace(
        minimal_input_data,
        classes=(
            replace(minimal_input_data.classes[0], class_name="まな"),
            *minimal_input_data.classes[1:],
        ),
        subjects=(
            replace(minimal_input_data.subjects[0], subject_name=""),
            *minimal_input_data.subjects[1:],
        ),
        teachers=(
            replace(minimal_input_data.teachers[0], teacher_name=""),
            *minimal_input_data.teachers[1:],
        ),
    )
    lesson = ScheduledLessonModel(
        "Q1",
        input_data.calendar_days[0].target_date,
        "P1",
        "T1",
        "R1",
        "C1",
        "CL1",
        "S1",
    )
    document = BuildTimetableDocumentService().execute(input_data, (lesson,))
    output = tmp_path / "blank-display-names.xlsx"

    ExcelTimetableWriterAdapter().write(document, output)

    worksheet = load_workbook(output, data_only=True)["全体"]
    assert worksheet["C4"].value == "まな"
    assert worksheet["C5"].value is None
    assert worksheet["C6"].value is None


def test_writer_places_dates_left_right_then_down_and_leaves_odd_slot_blank(
    minimal_input_data: InputDataModel,
    tmp_path: Path,
) -> None:
    document = BuildTimetableDocumentService().execute(minimal_input_data, ())
    output = tmp_path / "dates.xlsx"
    ExcelTimetableWriterAdapter().write(document, output)

    worksheet = load_workbook(output)["全体"]
    matrix_width = 2 + len(minimal_input_data.rooms)
    right_start = 1 + matrix_width + 1
    assert _excel_date(worksheet.cell(2, 1).value) == (
        minimal_input_data.calendar_days[0].target_date
    )
    assert (
        _excel_date(worksheet.cell(2, right_start).value)
        == minimal_input_data.calendar_days[1].target_date
    )
    assert _excel_date(worksheet.cell(24, 1).value) == (
        minimal_input_data.calendar_days[2].target_date
    )
    assert worksheet.cell(22, 1).value is None
    assert worksheet.cell(23, 1).value is None
    for row in range(24, 44):
        for column in range(right_start, right_start + matrix_width):
            cell = worksheet.cell(row, column)
            assert cell.value is None
            assert cell.border.left.style is None
            assert cell.border.right.style is None
    assert worksheet.cell(24, 1).value is not None
    assert worksheet.cell(26, 3).value is None


def test_writer_places_teacher_leaves_by_home_campus_and_formats_periods(
    minimal_input_data: InputDataModel,
    tmp_path: Path,
) -> None:
    input_data = replace(
        minimal_input_data,
        teachers=(
            minimal_input_data.teachers[0],
            replace(minimal_input_data.teachers[1], home_campus_id="C1"),
        ),
        teacher_leaves=(
            TeacherLeaveModel(
                "T1",
                minimal_input_data.calendar_days[0].target_date,
                tuple(period.period_id for period in minimal_input_data.periods),
            ),
            TeacherLeaveModel(
                "T2",
                minimal_input_data.calendar_days[0].target_date,
                tuple(period.period_id for period in minimal_input_data.periods),
            ),
            TeacherLeaveModel(
                "T1",
                minimal_input_data.calendar_days[1].target_date,
                ("P1", "P2"),
            ),
            TeacherLeaveModel(
                "T1",
                minimal_input_data.calendar_days[2].target_date,
                ("P3", "P4", "P5"),
            ),
        ),
    )
    document = BuildTimetableDocumentService().execute(input_data, ())
    output = tmp_path / "teacher-leaves.xlsx"

    writer = ExcelTimetableWriterAdapter()
    writer.write(document, output)

    worksheet = load_workbook(output)["全体"]
    matrix_width = 2 + len(minimal_input_data.rooms)
    right_start = 1 + matrix_width + 1
    assert worksheet["C22"].value == "教師一"
    assert worksheet["D22"].value == "教師二"
    assert worksheet.cell(22, right_start + 2).value == "教師一"
    assert worksheet.cell(22, right_start + 3).value == "①②"
    assert worksheet["C44"].value == "教師一"
    assert worksheet["D44"].value == "③～⑤"  # noqa: RUF001
    assert worksheet["C22"].border.left.style is None
    assert (worksheet["C22"].font.name, worksheet["C22"].font.sz) == (
        worksheet["C21"].font.name,
        worksheet["C21"].font.sz,
    )
    assert worksheet["C22"].alignment.horizontal == "center"
    assert worksheet.row_dimensions[22].height == worksheet.row_dimensions[21].height

    assert writer._format_teacher_leave_periods(("P1",), document) == "①"
    assert writer._format_teacher_leave_periods(("P1", "P2", "P5"), document) == "①②⑤"
    assert (
        writer._format_teacher_leave_periods(
            ("P1", "P2", "P3", "P5"),
            document,
        )
        == "①～③⑤"  # noqa: RUF001
    )


def test_document_rejects_teacher_leave_output_overflow(
    minimal_input_data: InputDataModel,
) -> None:
    input_data = replace(
        minimal_input_data,
        teachers=(
            minimal_input_data.teachers[0],
            replace(minimal_input_data.teachers[1], home_campus_id="C1"),
        ),
        teacher_leaves=(
            TeacherLeaveModel(
                "T1",
                minimal_input_data.calendar_days[0].target_date,
                ("P1",),
            ),
            TeacherLeaveModel(
                "T2",
                minimal_input_data.calendar_days[0].target_date,
                ("P2",),
            ),
        ),
    )

    try:
        BuildTimetableDocumentService().execute(input_data, ())
    except ValueError as error:
        assert "OUTPUT_TEACHER_LEAVE_OVERFLOW" in str(error)
    else:
        raise AssertionError("教師休みの出力領域超過が検出されませんでした")
