from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from school_timetable_solver.adapter.excel_input_adapter import ExcelInputReaderAdapter

SAMPLE = Path("projects/sample/input/時間割入力_サンプル.xlsx")


def test_reader_accepts_contract_v0_9_with_empty_optional_rows() -> None:
    result = ExcelInputReaderAdapter().read(SAMPLE)

    assert result.input_data is not None
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]
    assert result.input_data.settings.schema_version == "1.0"


def test_reader_reads_required_lesson_periods(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["11_配置ルール"]["T2"] = "P1|P2"
    target = tmp_path / "required-lesson-periods.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    assert result.input_data.placement_rules[0].required_lesson_period_ids == ("P1", "P2")
    assert len(result.input_data.periods) == 6
    assert result.input_data.teachers[0].home_campus_id == "C1"
    assert not result.input_data.teacher_leaves
    assert not result.input_data.lesson_count_rule_segments
    assert not result.input_data.lesson_count_preference_rule_segments
    assert not result.input_data.homeroom_boundary_rules
    room_priorities = {room.room_id: room.priority for room in result.input_data.rooms}
    assert room_priorities == {"R1": 0, "R2": 100, "R3": 100, "R4": 0}
    exam_rule = next(
        rule for rule in result.input_data.placement_rules if rule.rule_id == "RULE_ATTENDANCE_EXAM"
    )
    non_exam_rule = next(
        rule
        for rule in result.input_data.placement_rules
        if rule.rule_id == "RULE_ATTENDANCE_NON_EXAM"
    )
    assert exam_rule.attendance_streak_limit is None
    assert exam_rule.preferred_attendance_streak_limit == 3
    assert non_exam_rule.attendance_streak_limit == 3
    assert non_exam_rule.preferred_attendance_streak_limit == 2


def test_reader_normalizes_full_day_and_partial_teacher_leaves(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE)
    sheet = workbook["10_教師休み"]
    sheet.append(("T1", date(2026, 7, 27), "ALL", "終日休み"))
    sheet.append(("T2", date(2026, 7, 28), "P4|P6", "時限休み"))
    target = tmp_path / "teacher-leaves.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    teacher_leaves = result.input_data.teacher_leaves
    assert teacher_leaves[0].unavailable_period_ids == (
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
    )
    assert teacher_leaves[1].unavailable_period_ids == ("P4", "P6")


def test_reader_normalizes_whitespace_only_teacher_name_to_empty_string(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["06_教師"].append(
        ("T_BLANK", "   ", "C1", True, "空き教師"),
    )
    target = tmp_path / "blank-teacher-name.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    teacher = next(item for item in result.input_data.teachers if item.teacher_id == "T_BLANK")
    assert teacher.teacher_name == ""


def test_reader_normalizes_empty_subject_name_to_empty_string(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["08_教科"].append(
        ("S_BLANK", None, "other", True, "表示名が空白の教科"),
    )
    target = tmp_path / "blank-subject-name.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    subject = next(item for item in result.input_data.subjects if item.subject_id == "S_BLANK")
    assert subject.subject_name == ""


def test_reader_reads_lesson_count_rule_segments(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["13_授業配置数ルール"].append(
        (
            "LC1",
            "LC1_SEG1",
            "対象範囲に1コマ",
            True,
            "CL1",
            "S1",
            1,
            date(2026, 7, 27),
            date(2026, 7, 28),
            "P1|P3",
            None,
        )
    )
    target = tmp_path / "lesson-count-rules.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    segment = result.input_data.lesson_count_rule_segments[0]
    assert segment.rule_id == "LC1"
    assert segment.target_period_ids == ("P1", "P3")


def test_reader_reads_lesson_count_preference_rule_segments(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["14_授業配置数選好ルール"].append(
        (
            "LP1",
            "LP1_SEG1",
            "対象範囲を0コマに近づける",
            True,
            "CL1",
            "S1",
            0,
            date(2026, 7, 27),
            date(2026, 7, 28),
            "P3",
            None,
        )
    )
    target = tmp_path / "lesson-count-preference-rules.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    segment = result.input_data.lesson_count_preference_rule_segments[0]
    assert segment.rule_id == "LP1"
    assert segment.preferred_periods == 0
    assert segment.target_period_ids == ("P3",)


def test_reader_reads_teacher_day_off_rules(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["15_教師休日日数ルール"].append(
        (
            "DAY_OFF_T1_EARLY",
            "T1",
            True,
            "2026-07-27|2026-07-29",
            1,
            None,
            None,
            None,
            None,
            None,
            "期間内に終日休み1日",
        )
    )
    target = tmp_path / "teacher-day-off-rules.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    rule = result.input_data.teacher_day_off_rules[0]
    assert rule.rule_id == "DAY_OFF_T1_EARLY"
    assert rule.teacher_id == "T1"
    assert rule.required_days_off == 1
    assert rule.minimum_days_off is None
    assert rule.eligible_dates == (date(2026, 7, 27), date(2026, 7, 29))


def test_reader_reads_homeroom_boundary_rules(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["16_担任授業期間ルール"].append(
        (
            "HB1",
            "夏期前半",
            True,
            "division|has_regular_homeroom_lesson",
            "in|eq",
            "elementary/junior_high|TRUE",
            date(2026, 7, 27),
            date(2026, 7, 28),
            "担任の通常授業を期間端に配置",
        )
    )
    target = tmp_path / "homeroom-boundary-rules.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    rule = result.input_data.homeroom_boundary_rules[0]
    assert rule.rule_id == "HB1"
    assert rule.condition_fields == ("division", "has_regular_homeroom_lesson")
    assert rule.start_date == date(2026, 7, 27)


def test_operation_sheet_needs_no_header_and_preference_rows_are_ignored(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE)
    operation = workbook["00_操作説明"]
    operation.delete_rows(1, operation.max_row)
    operation["C8"] = "自由記述"
    preference = workbook["12_選好設定"]
    preference.append(("UNIMPLEMENTED", "not-a-boolean", "not-an-integer", "ignored"))
    target = tmp_path / "free-form.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is not None
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]


def test_reader_rejects_non_boolean_values(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["02_開講カレンダー"]["B2"] = 1
    target = tmp_path / "invalid-boolean.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is None
    assert any(issue.rule_id == "INVALID_BOOLEAN" for issue in result.issues)


def test_reader_rejects_all_mixed_with_individual_teacher_leave_periods(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["10_教師休み"].append(("T1", date(2026, 7, 27), "ALL|P1", None))
    target = tmp_path / "invalid-teacher-leave-periods.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is None
    assert any(issue.rule_id == "INVALID_TEACHER_LEAVE_PERIODS" for issue in result.issues)


def test_reader_rejects_string_date_and_unsupported_schema(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    workbook["02_開講カレンダー"]["A2"] = "2026-07-27"
    workbook["01_基本設定"]["B2"] = "9.9"
    target = tmp_path / "invalid-date-schema.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    rule_ids = {issue.rule_id for issue in result.issues}
    assert {"INVALID_DATE", "UNSUPPORTED_SCHEMA_VERSION"} <= rule_ids


def test_reader_trims_text_skips_empty_rows_and_uses_pipe_and_slash() -> None:
    result = ExcelInputReaderAdapter().read(SAMPLE)
    assert result.input_data is not None

    input_data = result.input_data
    elementary = next(rule for rule in input_data.placement_rules if rule.rule_id == "R_ELEMENTARY")
    override = next(rule for rule in input_data.placement_rules if rule.rule_id == "R_JH_OVERRIDE")
    assert elementary.condition_fields == ("division",)
    assert elementary.allowed_period_ids == ("P1", "P2", "P3")
    assert override.condition_values == ("junior_high", "1/3")


def test_missing_required_sheet_is_format_error(tmp_path: Path) -> None:
    workbook = load_workbook(SAMPLE)
    del workbook["14_授業配置数選好ルール"]
    target = tmp_path / "missing-sheet.xlsx"
    workbook.save(target)

    result = ExcelInputReaderAdapter().read(target)

    assert result.input_data is None
    assert any(issue.rule_id == "REQUIRED_SHEET_MISSING" for issue in result.issues)
