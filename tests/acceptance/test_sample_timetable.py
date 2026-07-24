from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from school_timetable_solver.main import main


def test_sample_cli_generates_one_sheet_verified_matrix_workbook(tmp_path: Path) -> None:
    output = tmp_path / "result.xlsx"
    log = tmp_path / "run.log"

    exit_code = main(
        (
            "--input",
            "projects/sample/input/時間割入力_サンプル.xlsx",
            "--output",
            str(output),
            "--log",
            str(log),
            "--mode",
            "strict",
            "--max-solve-seconds",
            "10",
            "--random-seed",
            "7",
            "--num-search-workers",
            "1",
        )
    )

    assert exit_code == 0
    assert output.is_file()
    assert log.is_file()
    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["全体"]
    worksheet = workbook["全体"]
    assert worksheet["A2"].value is not None
    assert worksheet["G2"].value is not None
    assert worksheet["A24"].value is not None
    assert worksheet["G24"].value is None
    assert {"A2:B2", "A3:B3", "C2:D2", "A4:A6"} <= {
        str(item) for item in worksheet.merged_cells.ranges
    }
    class_rows = (4, 7, 10, 13, 16, 19, 26, 29, 32, 35, 38, 41)
    generated_class_cells = [
        worksheet.cell(row, column).value
        for row in class_rows
        for column in (3, 4, 5, 9, 10, 11)
        if worksheet.cell(row, column).value is not None
    ]
    assert len(generated_class_cells) == 4
    assert worksheet["A2"].border.top.style == "thin"
    assert worksheet["C4"].border.bottom.style == "hair"
    assert workbook.__dict__.get("_external_links") == []


def test_validate_only_runs_without_solver_or_workbook_output(tmp_path: Path) -> None:
    output = tmp_path / "validation.xlsx"
    log = tmp_path / "validate.log"

    exit_code = main(
        (
            "--input",
            "projects/sample/input/時間割入力_サンプル.xlsx",
            "--output",
            str(output),
            "--log",
            str(log),
            "--mode",
            "validate_only",
        )
    )

    assert exit_code == 0
    assert not output.exists()
    assert log.is_file()
    assert "Solver完了" not in log.read_text(encoding="utf-8")


def test_input_error_does_not_replace_existing_output(tmp_path: Path) -> None:
    output = tmp_path / "existing.xlsx"
    output.write_bytes(b"keep-me")

    exit_code = main(
        (
            "--input",
            str(tmp_path / "missing.xlsx"),
            "--output",
            str(output),
        )
    )

    assert exit_code == 2
    assert output.read_bytes() == b"keep-me"
