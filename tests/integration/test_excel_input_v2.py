from __future__ import annotations

from dataclasses import replace

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from school_timetable_solver.adapter.excel_input_router import CompatibleExcelInputReaderAdapter
from school_timetable_solver.adapter.excel_input_v2_adapter import BLANK_TOKEN, NO, YES
from school_timetable_solver.adapter.excel_v2_workbook_adapter import ExcelV2WorkbookWriterAdapter
from school_timetable_solver.adapter.excel_v2_workbook_postprocessor import (
    ExcelV2WorkbookPostprocessor,
)
from school_timetable_solver.model.input_models import InputDataModel, PlacementRuleModel
from school_timetable_solver.model.master_models import SubjectModel, TeacherModel
from school_timetable_solver.service.planning_services import RuleResolverService


def test_v2_round_trip_preserves_input_semantics(tmp_path, minimal_input_data: InputDataModel) -> None:
    path = tmp_path / "input_v2.xlsx"
    ExcelV2WorkbookWriterAdapter().write(path, minimal_input_data)

    result = CompatibleExcelInputReaderAdapter().read(path)

    assert result.input_data is not None, result.issues
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]
    actual = result.input_data
    expected = minimal_input_data

    assert actual.settings.schema_version == "2.0"
    assert actual.settings.timetable_name == expected.settings.timetable_name
    assert actual.calendar_days == expected.calendar_days
    assert actual.periods == expected.periods
    assert actual.campuses == expected.campuses
    assert actual.rooms == expected.rooms
    assert actual.teachers == expected.teachers
    assert actual.classes == expected.classes
    assert actual.subjects == expected.subjects
    assert actual.lesson_requirements == expected.lesson_requirements
    assert actual.teacher_leaves == expected.teacher_leaves
    assert actual.placement_rules == expected.placement_rules
    assert actual.lesson_count_rule_segments == expected.lesson_count_rule_segments
    assert actual.lesson_count_preference_rule_segments == expected.lesson_count_preference_rule_segments
    assert actual.teacher_day_off_rules == expected.teacher_day_off_rules
    assert actual.homeroom_boundary_rules == expected.homeroom_boundary_rules
    assert actual.class_pair_overlap_rules == expected.class_pair_overlap_rules


def test_v2_workbook_is_self_describing_and_uses_human_tokens(
    tmp_path,
    minimal_input_data: InputDataModel,
) -> None:
    path = tmp_path / "input_v2.xlsx"
    ExcelV2WorkbookWriterAdapter().write(path, minimal_input_data)

    workbook = load_workbook(path, data_only=False)
    try:
        assert "00_最初に読む" in workbook.sheetnames
        assert "90_AI編集ガイド" in workbook.sheetnames
        assert "08_入力確認" in workbook.sheetnames
        assert "_system" in workbook.sheetnames
        assert workbook["_system"].sheet_state == "hidden"

        intro_text = "\n".join(
            str(cell.value)
            for row in workbook["00_最初に読む"].iter_rows()
            for cell in row
            if cell.value is not None
        )
        guide_text = "\n".join(
            str(cell.value)
            for row in workbook["90_AI編集ガイド"].iter_rows()
            for cell in row
            if cell.value is not None
        )
        calendar_values = {
            str(cell.value)
            for row in workbook["01_日程"].iter_rows()
            for cell in row
            if cell.value in {YES, NO}
        }

        assert BLANK_TOKEN in intro_text
        assert "日付集合" in guide_text
        assert "一般化" in guide_text
        assert calendar_values == {YES, NO}
    finally:
        workbook.close()


def test_v2_migration_preserves_resolved_meaning_for_direct_class_rule_without_campus(
    tmp_path,
    minimal_input_data: InputDataModel,
) -> None:
    target_class = minimal_input_data.classes[0]
    rule = PlacementRuleModel(
        rule_id="RULE_DIRECT_CLASS",
        rule_name="特定クラス時限制限",
        enabled=True,
        constraint_type="hard",
        target_entity="class",
        condition_fields=("class_id",),
        condition_operators=("eq",),
        condition_values=(target_class.class_id,),
        campus_id=None,
        start_date=None,
        end_date=None,
        weekdays=(),
        allowed_period_ids=("P1", "P2"),
        daily_hard_limit=None,
        forbid_first_last_same_day=None,
        attendance_streak_limit=None,
        priority=50,
    )
    source = replace(minimal_input_data, placement_rules=(*minimal_input_data.placement_rules, rule))
    path = tmp_path / "input_v2.xlsx"

    ExcelV2WorkbookWriterAdapter().write(path, source)
    ExcelV2WorkbookPostprocessor().execute(path, source)
    result = CompatibleExcelInputReaderAdapter().read(path)

    assert result.input_data is not None, result.issues
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]
    expected_resolved = RuleResolverService().execute(source)
    actual_resolved = RuleResolverService().execute(result.input_data)
    expected_target = [
        item
        for item in expected_resolved.class_date_rules
        if item.class_id == target_class.class_id
    ]
    actual_target = [
        item
        for item in actual_resolved.class_date_rules
        if item.class_id == target_class.class_id
    ]
    assert [item.allowed_period_ids for item in actual_target] == [
        item.allowed_period_ids for item in expected_target
    ]

    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["06_基本配置ルール"]
        assert not worksheet.column_dimensions["A"].hidden
    finally:
        workbook.close()


def test_v2_reference_labels_separate_input_identity_from_blank_output_name(
    tmp_path,
    minimal_input_data: InputDataModel,
) -> None:
    campus_id = minimal_input_data.campuses[0].campus_id
    blank_teachers = (
        TeacherModel("TEACHER_BLANK_A", "", campus_id, True),
        TeacherModel("TEACHER_BLANK_B", "", campus_id, True),
    )
    blank_subjects = (
        SubjectModel("SUBJECT_BLANK_A", "", "other", True),
        SubjectModel("SUBJECT_BLANK_B", "", "other", True),
    )
    source = replace(
        minimal_input_data,
        teachers=(*minimal_input_data.teachers, *blank_teachers),
        subjects=(*minimal_input_data.subjects, *blank_subjects),
    )
    path = tmp_path / "input_v2_reference_labels.xlsx"

    ExcelV2WorkbookWriterAdapter().write(path, source)
    ExcelV2WorkbookPostprocessor().execute(path, source)
    result = CompatibleExcelInputReaderAdapter().read(path)

    assert result.input_data is not None, result.issues
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]
    assert result.input_data.teachers == source.teachers
    assert result.input_data.subjects == source.subjects

    workbook = load_workbook(path, data_only=False)
    try:
        system_sheet = workbook["_system"]
        assert "T_REFERENCE_MAP" in system_sheet.tables
        table = system_sheet.tables["T_REFERENCE_MAP"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            system_sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        rows = [
            {
                str(header): system_sheet.cell(row, column).value
                for header, column in zip(headers, range(min_col, max_col + 1), strict=True)
            }
            for row in range(min_row + 1, max_row + 1)
        ]
        teacher_rows = [row for row in rows if row["参照種別"] == "teacher"]
        subject_rows = [row for row in rows if row["参照種別"] == "subject"]
        assert len({row["入力用ラベル"] for row in teacher_rows}) == len(teacher_rows)
        assert len({row["入力用ラベル"] for row in subject_rows}) == len(subject_rows)
        assert {
            row["内部ID"]: row["出力表示名"]
            for row in teacher_rows
            if row["内部ID"] in {item.teacher_id for item in blank_teachers}
        } == {"TEACHER_BLANK_A": None, "TEACHER_BLANK_B": None}
        assert {
            row["内部ID"]: row["出力表示名"]
            for row in subject_rows
            if row["内部ID"] in {item.subject_id for item in blank_subjects}
        } == {"SUBJECT_BLANK_A": None, "SUBJECT_BLANK_B": None}
    finally:
        workbook.close()
