from __future__ import annotations

from dataclasses import replace

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from school_timetable_solver.adapter.excel_input_router import CompatibleExcelInputReaderAdapter
from school_timetable_solver.adapter.excel_v2_reference_postprocessor import (
    ReferenceLabelExcelV2WorkbookPostprocessor,
)
from school_timetable_solver.adapter.excel_v2_workbook_adapter import ExcelV2WorkbookWriterAdapter
from school_timetable_solver.model.input_models import ClassPairOverlapRuleModel, InputDataModel


def test_v2_reference_labels_distinguish_same_class_output_name_within_campus(
    tmp_path,
    minimal_input_data: InputDataModel,
) -> None:
    ordinary = minimal_input_data.classes[0]
    special = replace(
        ordinary,
        class_id="CLASS_SAME_NAME_SPECIAL",
        exam_category="special",
        homeroom_teacher_id=None,
    )
    ordinary_requirement = next(
        item
        for item in minimal_input_data.lesson_requirements
        if item.class_id == ordinary.class_id
    )
    special_requirement = replace(
        ordinary_requirement,
        requirement_id="REQ_SAME_NAME_SPECIAL",
        class_id=special.class_id,
    )
    pair = ClassPairOverlapRuleModel(
        rule_id="PAIR_SAME_NAME",
        rule_name="同名クラス重複禁止",
        enabled=True,
        first_class_id=ordinary.class_id,
        second_class_id=special.class_id,
    )
    source = replace(
        minimal_input_data,
        classes=(*minimal_input_data.classes, special),
        lesson_requirements=(*minimal_input_data.lesson_requirements, special_requirement),
        class_pair_overlap_rules=(*minimal_input_data.class_pair_overlap_rules, pair),
    )
    path = tmp_path / "input_v2_same_class_name.xlsx"

    ExcelV2WorkbookWriterAdapter().write(path, source)
    ReferenceLabelExcelV2WorkbookPostprocessor().execute(path, source)
    result = CompatibleExcelInputReaderAdapter().read(path)

    assert result.input_data is not None, result.issues
    assert not [issue for issue in result.issues if issue.severity == "ERROR"]
    assert result.input_data.classes == source.classes
    assert result.input_data.lesson_requirements == source.lesson_requirements
    assert result.input_data.class_pair_overlap_rules == source.class_pair_overlap_rules

    workbook = load_workbook(path, data_only=False)
    try:
        system_sheet = workbook["_system"]
        table = system_sheet.tables["T_REFERENCE_MAP"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            system_sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        rows = [
            {
                str(header): system_sheet.cell(row, column).value
                for header, column in zip(
                    headers,
                    range(min_col, max_col + 1),
                    strict=True,
                )
            }
            for row in range(min_row + 1, max_row + 1)
        ]
        class_rows = {
            row["内部ID"]: row
            for row in rows
            if row["参照種別"] == "class"
            and row["内部ID"] in {ordinary.class_id, special.class_id}
        }
        assert class_rows[ordinary.class_id]["出力表示名"] == ordinary.class_name
        assert class_rows[special.class_id]["出力表示名"] == ordinary.class_name
        assert (
            class_rows[ordinary.class_id]["入力用ラベル"]
            != class_rows[special.class_id]["入力用ラベル"]
        )
    finally:
        workbook.close()
